from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import math
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook

try:
    from .ai_generation import generate_ai_insights, load_prompt
    from .fiscal_history import apple_period, build_history_context
    from .insight_evidence import build_evidence_packet
except ImportError:  # pragma: no cover - keeps direct script execution working.
    from ai_generation import generate_ai_insights, load_prompt
    from fiscal_history import apple_period, build_history_context
    from insight_evidence import build_evidence_packet


MONTHLY_ACCOUNT_SHEET = "\u5c0f\u7ea2\u4e66\u603b\u6570\u636e"
MONTHLY_NOTE_SHEET = "\u5c0f\u7ea2\u4e66\u7b14\u8bb0\u6570\u636e\u5e93"
KPI_QUARTER = "Q4"
LOGIC_VERSION = "2026-08-08.1"


class SourceDiscoveryError(RuntimeError):
    """Raised when the required monthly source workbooks are not unambiguous."""


def validate_month(month: str) -> str:
    if not isinstance(month, str) or not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("Month must use YYYY-MM format")
    try:
        datetime.strptime(month, "%Y-%m")
    except ValueError as error:
        raise ValueError("Month must use YYYY-MM format") from error
    return month


def previous_month(month: str) -> str:
    current = datetime.strptime(validate_month(month), "%Y-%m")
    if current.month == 1:
        return f"{current.year - 1}-12"
    return f"{current.year}-{current.month - 1:02d}"


def _is_curated_monthly_workbook(path: Path) -> bool:
    name = path.name.lower()
    if "errydoc" in name or "template" in name or "\u6a21\u677f" in path.name:
        return False
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return {MONTHLY_ACCOUNT_SHEET, MONTHLY_NOTE_SHEET}.issubset(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception:
        return False


def q4_months_to_date(month: str) -> list[str]:
    current = datetime.strptime(validate_month(month), "%Y-%m")
    if current.month not in {7, 8, 9}:
        return [month]
    return [f"{current.year}-{month_number:02d}" for month_number in range(7, current.month + 1)]


def _discover_monthly_workbook(root: Path, month: str) -> Path:
    monthly_dir = root / f"{month}\u722c\u866b\u6570\u636e"
    candidates = sorted(
        (path for path in monthly_dir.glob("*.xlsx") if path.is_file()),
        key=lambda path: path.name.lower(),
    ) if monthly_dir.is_dir() else []
    curated = [path for path in candidates if _is_curated_monthly_workbook(path)]
    candidate_list = ", ".join(str(path.resolve()) for path in candidates) or "(none)"
    if not curated:
        raise SourceDiscoveryError(
            f"No valid curated monthly workbook for {month}; candidates: {candidate_list}"
        )
    if len(curated) != 1:
        curated_list = ", ".join(str(path.resolve()) for path in curated)
        raise SourceDiscoveryError(
            f"Ambiguous curated monthly workbooks for {month}: {curated_list}"
        )
    return curated[0].resolve()


def discover_month_sources(data_root: Path, month: str) -> dict[str, Any]:
    validated_month = validate_month(month)
    root = Path(data_root).expanduser().resolve()
    q4_actual_workbooks = [
        {"month": required_month, "path": _discover_monthly_workbook(root, required_month)}
        for required_month in q4_months_to_date(validated_month)
    ]

    fiscal_year = validated_month[2:4]
    kpi_workbook = root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI" / f"FY{fiscal_year}.xlsx"
    if not kpi_workbook.is_file():
        raise SourceDiscoveryError(f"KPI workbook not found: {kpi_workbook.resolve()}")
    return {
        "monthly_workbook": next(
            item["path"] for item in q4_actual_workbooks if item["month"] == validated_month
        ),
        "kpi_workbook": kpi_workbook.resolve(),
        "q4_actual_workbooks": q4_actual_workbooks,
    }


def _find_header_row(worksheet: Any, required_fields: dict[str, tuple[str, ...]]) -> tuple[int, list[str]]:
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        headers = [normalize_text(value) for value in row]
        if all(any(alias in headers for alias in aliases) for aliases in required_fields.values()):
            return row_number, headers
        if row_number >= 50:
            break
    labels = ", ".join(required_fields)
    raise ValueError(f"Could not find a header row containing: {labels}")


def _header_index(headers: list[str], aliases: tuple[str, ...], start_at: int = 0) -> int:
    for index, header in enumerate(headers[start_at:], start=start_at):
        if header in aliases:
            return index
    raise ValueError(f"Missing required column: {' / '.join(aliases)}")


def _identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_rows(path: Path, sheet_name: str, required_fields: dict[str, tuple[str, ...]]) -> tuple[Any, int, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Sheet not found in {path}: {sheet_name}")
    worksheet = workbook[sheet_name]
    try:
        header_row, headers = _find_header_row(worksheet, required_fields)
    except Exception:
        workbook.close()
        raise
    return workbook, header_row, headers


def read_monthly_accounts(path: Path, sheet_name: str, month: str) -> list[dict]:
    required = {
        "dealer": ("\u7ecf\u9500\u5546\u540d\u79f0",),
        "store": ("\u95e8\u5e97\u540d\u79f0",),
        "account_name": ("\u5c0f\u7ea2\u4e66\u8d26\u53f7\u540d\u79f0",),
        "author_id": ("\u5c0f\u7ea2\u4e66\u4f5c\u8005ID",),
        "xhs_id": ("\u5c0f\u7ea2\u4e66\u53f7",),
        "reads": ("\u603b\u6d4f\u89c8",),
        "new_fans": ("\u65b0\u589e\u7c89\u4e1d",),
        "likes": ("\u70b9\u8d5e",),
        "collects": ("\u6536\u85cf",),
        "comments": ("\u8bc4\u8bba",),
        "visitors": ("\u4e3b\u9875\u8bbf\u5ba2\u6570",),
    }
    workbook, header_row, headers = _read_rows(Path(path), sheet_name, required)
    try:
        identifiers_end = max(
            _header_index(headers, required[field])
            for field in ("dealer", "store", "account_name", "author_id", "xhs_id")
        )
        indexes = {
            field: _header_index(headers, aliases, identifiers_end + 1 if field in {"reads", "new_fans", "likes", "collects", "comments", "visitors"} else 0)
            for field, aliases in required.items()
        }
        records = []
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            author_id = _identifier(row[indexes["author_id"]] if len(row) > indexes["author_id"] else None)
            if not author_id:
                continue
            record = {
                "source_month": validate_month(month),
                "dealer": _identifier(row[indexes["dealer"]] if len(row) > indexes["dealer"] else None),
                "store": _identifier(row[indexes["store"]] if len(row) > indexes["store"] else None),
                "account_name": _identifier(row[indexes["account_name"]] if len(row) > indexes["account_name"] else None),
                "author_id": author_id,
                "xhs_id": _identifier(row[indexes["xhs_id"]] if len(row) > indexes["xhs_id"] else None),
            }
            record.update({
                field: safe_number(row[index] if len(row) > index else None)
                for field, index in indexes.items()
                if field in {"reads", "new_fans", "likes", "collects", "comments", "visitors"}
            })
            records.append(record)
        return records
    finally:
        workbook.close()


def aggregate_q4_actuals(monthly_accounts: list[tuple[str, list[dict]]]) -> dict[str, dict]:
    actuals: dict[str, dict] = {}
    for month, accounts in monthly_accounts:
        validated_month = validate_month(month)
        for account in accounts:
            author_id = _identifier(account.get("author_id"))
            if not author_id:
                continue
            item = actuals.setdefault(author_id, {
                "author_id": author_id,
                "source_months": [],
                "reads": 0.0,
                "new_fans": 0.0,
                "likes": 0.0,
                "collects": 0.0,
                "comments": 0.0,
            })
            if validated_month not in item["source_months"]:
                item["source_months"].append(validated_month)
            for metric in ("reads", "new_fans", "likes", "collects", "comments"):
                item[metric] += safe_number(account.get(metric))
    return actuals


def read_notes(path: Path, sheet_name: str, snapshot_date: str) -> list[dict]:
    required = {
        "note_format": ("\u7b14\u8bb0\u5f62\u5f0f", "\u7b14\u8bb0\u7c7b\u578b"),
        "category": ("\u7b14\u8bb0\u7c7b\u578b", "\u7b14\u8bb0\u5f62\u5f0f"),
        "exported_at": ("\u65f6\u95f4", "\u5bfc\u51fa\u6570\u636e\u65f6\u95f4"),
        "note_id": ("\u7b14\u8bb0ID",),
        "author_id": ("\u4f5c\u8005ID",),
        "reads": ("\u9605\u8bfb\u6b21\u6570", "\u7d2f\u8ba1\u9605\u8bfb\u6570"),
        "likes": ("\u70b9\u8d5e\u6b21\u6570", "\u7d2f\u8ba1\u70b9\u8d5e\u6570"),
        "collects": ("\u6536\u85cf\u6b21\u6570", "\u7d2f\u8ba1\u6536\u85cf\u6570"),
        "comments": ("\u8bc4\u8bba\u6b21\u6570", "\u7d2f\u8ba1\u8bc4\u8bba\u6570"),
        "shares": ("\u5206\u4eab\u6b21\u6570", "\u7d2f\u8ba1\u5206\u4eab\u6570"),
        "new_fans": ("\u5355\u6761\u7b14\u8bb0\u6da8\u7c89\u6570", "\u7b14\u8bb0\u65b0\u589e\u7c89\u4e1d\u6570"),
    }
    published_aliases = ("\u7b14\u8bb0\u53d1\u5e03\u65f6\u95f4", "\u53d1\u5e03\u65f6\u95f4", "\u53d1\u5e03\u65e5\u671f")
    workbook, header_row, headers = _read_rows(Path(path), sheet_name, required)
    try:
        indexes = {field: _header_index(headers, aliases) for field, aliases in required.items()}
        publish_index = next((index for index, header in enumerate(headers) if header in published_aliases), None)
        if "\u7b14\u8bb0\u5f62\u5f0f" in headers and "\u7b14\u8bb0\u7c7b\u578b" in headers:
            indexes["note_format"] = _header_index(headers, ("\u7b14\u8bb0\u5f62\u5f0f",))
            indexes["category"] = _header_index(headers, ("\u7b14\u8bb0\u7c7b\u578b",))
        if indexes["note_format"] == indexes["category"]:
            raise ValueError("Note format and category must use distinct source columns")
        records = []
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            note_id = _identifier(row[indexes["note_id"]] if len(row) > indexes["note_id"] else None)
            if not note_id:
                continue
            record = {
                "snapshot_date": _identifier(snapshot_date),
                "note_id": note_id,
                "author_id": _identifier(row[indexes["author_id"]] if len(row) > indexes["author_id"] else None),
                "note_format": _identifier(row[indexes["note_format"]] if len(row) > indexes["note_format"] else None),
                "category": _identifier(row[indexes["category"]] if len(row) > indexes["category"] else None),
                "exported_at": _identifier(row[indexes["exported_at"]] if len(row) > indexes["exported_at"] else None),
                "publish_date": _identifier(row[publish_index] if publish_index is not None and len(row) > publish_index else None),
            }
            record.update({
                field: safe_number(row[index] if len(row) > index else None)
                for field, index in indexes.items()
                if field in {"reads", "likes", "collects", "comments", "shares", "new_fans"}
            })
            records.append(record)
        return records
    finally:
        workbook.close()


def read_kpi(path: Path, quarter: str) -> list[dict]:
    quarter_name = normalize_text(quarter).upper()
    if not re.fullmatch(r"Q[1-4]", quarter_name):
        raise ValueError("Quarter must use Q1 through Q4")
    required = {
        "group": ("\u7ecf\u9500\u5546\u5206\u7ec4",),
        "account_name": ("\u7ecf\u9500\u5546\u5c0f\u7ea2\u4e66\u8d26\u53f7",),
        "system_account_name": ("\u7cfb\u7edf\u8d26\u53f7\u540d\u79f0",),
        "author_id": ("\u8d26\u53f7ID",),
        "read_target": (f"{quarter_name} \u603b\u9605\u8bfb\u91cf\u76ee\u6807",),
        "interaction_target": (f"{quarter_name} \u603b\u4e92\u52a8\u91cf\u76ee\u6807",),
        "fan_target": (f"{quarter_name} \u603b\u65b0\u589e\u7c89\u4e1d\u76ee\u6807",),
    }
    workbook, header_row, headers = _read_rows(Path(path), quarter_name, required)
    try:
        indexes = {field: _header_index(headers, aliases) for field, aliases in required.items()}
        records = []
        current_group = ""
        worksheet = workbook[quarter_name]
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            author_id = _identifier(row[indexes["author_id"]] if len(row) > indexes["author_id"] else None)
            if not author_id:
                continue
            group_value = _identifier(row[indexes["group"]] if len(row) > indexes["group"] else None)
            current_group = group_value or current_group
            records.append({
                "group": current_group,
                "account_name": _identifier(row[indexes["account_name"]] if len(row) > indexes["account_name"] else None),
                "system_account_name": _identifier(row[indexes["system_account_name"]] if len(row) > indexes["system_account_name"] else None),
                "author_id": author_id,
                "read_target": safe_number(row[indexes["read_target"]] if len(row) > indexes["read_target"] else None),
                "interaction_target": safe_number(row[indexes["interaction_target"]] if len(row) > indexes["interaction_target"] else None),
                "fan_target": safe_number(row[indexes["fan_target"]] if len(row) > indexes["fan_target"] else None),
            })
        return records
    finally:
        workbook.close()


def _region_overrides() -> dict:
    path = Path(__file__).resolve().parents[1] / "config" / "region_overrides.json"
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def _category_mapping() -> dict:
    path = Path(__file__).resolve().parents[1] / "config" / "category_mapping.json"
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def _serialize_source_files(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value.resolve())
    if isinstance(value, dict):
        return {key: _serialize_source_files(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_serialize_source_files(item) for item in value]
    return value


def derive_data_freshness(notes: list[dict], source_month: str) -> dict:
    candidates = []
    for note in notes:
        raw_value = note.get("exported_at")
        text_value = _identifier(raw_value)
        if not text_value:
            continue
        try:
            parsed = raw_value if isinstance(raw_value, datetime) else datetime.fromisoformat(text_value.replace("Z", "+00:00"))
        except ValueError:
            continue
        comparable = parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=timezone.utc)
        candidates.append((comparable.astimezone(timezone.utc), text_value))
    if candidates:
        _, source_snapshot_at = max(candidates, key=lambda item: (item[0], item[1]))
        return {
            "source_snapshot_at": source_snapshot_at,
            "basis": "note_export_timestamp",
            "is_fallback": False,
        }
    current = datetime.strptime(validate_month(source_month), "%Y-%m")
    month_end = calendar.monthrange(current.year, current.month)[1]
    return {
        "source_snapshot_at": f"{source_month}-{month_end:02d}",
        "basis": "source_month_end_fallback",
        "is_fallback": True,
    }


def profile_quality(
    accounts: list[dict],
    notes: list[dict],
    kpis: list[dict],
    source_month: str,
    source_files: dict[str, Path],
) -> dict:
    month = validate_month(source_month)
    account_ids = {normalize_text(record.get("author_id")) for record in accounts if normalize_text(record.get("author_id"))}
    kpi_ids = {normalize_text(record.get("author_id")) for record in kpis if normalize_text(record.get("author_id"))}
    note_ids = [normalize_text(record.get("note_id")) for record in notes if normalize_text(record.get("note_id"))]
    note_id_counts = Counter(note_ids)
    duplicate_note_ids = sorted(note_id for note_id, count in note_id_counts.items() if count > 1)
    matched_kpi_ids = kpi_ids & account_ids
    unmatched_kpi = [record for record in kpis if normalize_text(record.get("author_id")) not in account_ids]
    missing_categories = sum(not normalize_text(record.get("category")) for record in notes)

    overrides = _region_overrides()
    unknown_cities = sum(
        resolve_region(
            record.get("dealer"), record.get("store"), record.get("account_name"), overrides, record.get("author_id")
        )["confidence"] == "unknown"
        for record in accounts
    )
    cohort_counts = Counter(classify_account_cohort(record.get("author_id"), kpi_ids) for record in accounts)
    warnings = []
    if account_ids != kpi_ids:
        warnings.append({"code": "scope_change", "message": "Account and KPI source scopes differ."})
    if duplicate_note_ids:
        warnings.append({"code": "duplicate_note_ids", "message": f"Duplicate note IDs: {len(duplicate_note_ids)}."})
    if missing_categories:
        warnings.append({"code": "missing_categories", "message": f"Notes missing categories: {missing_categories}."})
    if unknown_cities:
        warnings.append({"code": "unknown_cities", "message": f"Accounts with unknown cities: {unknown_cities}."})
    if unmatched_kpi:
        warnings.append({"code": "unmatched_kpi_ids", "message": f"Unmatched KPI IDs: {len(unmatched_kpi)}."})

    errors = []
    if not accounts or not account_ids:
        errors.append("No account rows with author IDs were read.")
    if not notes or not note_ids:
        errors.append("No note rows with note IDs were read.")
    if not kpis or not kpi_ids:
        errors.append("No KPI rows with account IDs were read.")
    source_file_paths = _serialize_source_files(source_files)
    kpi_count = len(kpis)
    return {
        "source_month": month,
        "previous_month": previous_month(month),
        "source_files": source_file_paths,
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "data_freshness": derive_data_freshness(notes, month),
        "account_rows": len(accounts),
        "unique_author_ids": len(account_ids),
        "note_rows": len(notes),
        "unique_note_ids": len(note_id_counts),
        "duplicate_note_ids": duplicate_note_ids,
        "kpi_accounts": kpi_count,
        "matched_kpi_accounts": len(matched_kpi_ids),
        "unmatched_kpi_ids": sorted(normalize_text(record.get("author_id")) for record in unmatched_kpi),
        "unmatched_kpi_account_names": sorted(normalize_text(record.get("account_name")) for record in unmatched_kpi),
        "unmatched_account_ids": sorted(account_ids - kpi_ids),
        "unmatched_account_names": sorted(
            normalize_text(record.get("account_name"))
            for record in accounts
            if normalize_text(record.get("author_id")) not in kpi_ids
        ),
        "kpi_match_rate": len(matched_kpi_ids) / kpi_count if kpi_count else None,
        "category_completeness": {
            "present": len(notes) - missing_categories,
            "missing": missing_categories,
            "rate": (len(notes) - missing_categories) / len(notes) if notes else None,
        },
        "city_identification": {
            "identified": len(accounts) - unknown_cities,
            "unknown": unknown_cities,
            "rate": (len(accounts) - unknown_cities) / len(accounts) if accounts else None,
        },
        "account_cohorts": {
            "core_kpi": cohort_counts["core_kpi"],
            "expanded_store": cohort_counts["expanded_store"],
        },
        "warnings": warnings,
        "errors": errors,
        "publishable": not errors,
        "quality_status": "failed" if errors else ("ready_with_warnings" if warnings else "ready"),
    }


def _default_data_root() -> Path:
    return Path(__file__).resolve().parents[3] / "\u6570\u636e"


def _write_json(path: Path, value: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(value, handle, ensure_ascii=False, indent=2, allow_nan=False)
        handle.write("\n")


def _scoped_quality(payload: dict) -> dict:
    quality = payload.get("quality") or {}
    return {
        "quality_status": quality.get("quality_status", "ready"),
        "errors": list(quality.get("errors") or []),
    }


def build_dealer_index(payload: dict) -> dict:
    return {
        "schema_version": payload.get("schema_version"),
        "source_month": payload.get("source_month"),
        "generated_at": payload.get("generated_at"),
        "data_freshness": (payload.get("metadata") or {}).get("data_freshness"),
        "quality": _scoped_quality(payload),
        "dealers": [
            {"dealer_id": dealer.get("dealer_id"), "name": dealer.get("name")}
            for dealer in payload.get("dealers", [])
        ],
    }


def build_dealer_payload(payload: dict, dealer_id: str) -> dict:
    dealer = next(
        (item for item in payload.get("dealers", []) if item.get("dealer_id") == dealer_id),
        None,
    )
    if dealer is None:
        raise ValueError(f"Unknown dealer_id: {dealer_id}")
    return {
        "schema_version": payload.get("schema_version"),
        "source_month": payload.get("source_month"),
        "period": payload.get("period"),
        "generated_at": payload.get("generated_at"),
        "data_freshness": (payload.get("metadata") or {}).get("data_freshness"),
        "quality": _scoped_quality(payload),
        "history": payload.get("history"),
        "dealer": dealer,
    }


def build_apple_month_payload(payload: dict) -> dict:
    return {
        "schema_version": payload.get("schema_version"),
        "source_month": payload.get("source_month"),
        "period": payload.get("period"),
        "generated_at": payload.get("generated_at"),
        "metadata": payload.get("metadata"),
        "source_files": payload.get("source_files", {}),
        "quality": _scoped_quality(payload),
        "history": payload.get("history"),
        "apple": payload.get("apple"),
    }


def write_dealer_scoped_artifacts(output_dir: Path, payload: dict) -> None:
    output_dir = Path(output_dir)
    dealer_ids = [dealer.get("dealer_id") for dealer in payload.get("dealers", [])]
    for dealer_id in dealer_ids:
        if not isinstance(dealer_id, str) or not re.fullmatch(r"[a-z0-9-]+", dealer_id):
            raise ValueError(f"Unsafe dealer_id for artifact path: {dealer_id}")

    dealer_dir = output_dir / "dealers"
    dealer_dir.mkdir(parents=True, exist_ok=True)
    expected_files = {f"{dealer_id}.json" for dealer_id in dealer_ids}
    for existing_path in dealer_dir.glob("*.json"):
        if existing_path.name not in expected_files:
            existing_path.unlink()

    _write_json(output_dir / "dealer_index.json", build_dealer_index(payload))
    for dealer_id in dealer_ids:
        _write_json(output_dir / "dealers" / f"{dealer_id}.json", build_dealer_payload(payload, dealer_id))


def load_month_index(path: Path) -> dict:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return {"schema_version": "2.0", "months": []}
    if not isinstance(value, dict):
        return {"schema_version": "2.0", "months": []}
    value.setdefault("months", [])
    return value


def write_versioned_artifacts(output_dir: Path, payload: dict) -> dict[str, dict]:
    output_dir = Path(output_dir)
    month = payload["source_month"]
    month_dir = output_dir / "months" / month
    dealer_payloads = {
        dealer["dealer_id"]: build_dealer_payload(payload, dealer["dealer_id"])
        for dealer in payload.get("dealers", [])
        if isinstance(dealer.get("dealer_id"), str)
    }
    _write_json(month_dir / "apple.json", build_apple_month_payload(payload))
    for dealer_id, dealer_payload in dealer_payloads.items():
        if not re.fullmatch(r"[a-z0-9-]+", dealer_id):
            raise ValueError(f"Unsafe dealer_id for artifact path: {dealer_id}")
        _write_json(month_dir / "dealers" / f"{dealer_id}.json", dealer_payload)

    index_path = output_dir / "month_index.json"
    index = load_month_index(index_path)
    months = sorted(set(index.get("months") or []) | {month})
    index.update({
        "schema_version": "2.0",
        "latest_month": max(months),
        "months": months,
    })
    _write_json(index_path, index)
    return dealer_payloads


def _failed_quality_report(month: str, error: Exception) -> dict:
    try:
        data_freshness = derive_data_freshness([], month)
    except ValueError:
        data_freshness = {"source_snapshot_at": None, "basis": "unavailable", "is_fallback": True}
    return {
        "source_month": month,
        "source_files": {},
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "data_freshness": data_freshness,
        "account_rows": 0,
        "unique_author_ids": 0,
        "note_rows": 0,
        "unique_note_ids": 0,
        "kpi_accounts": 0,
        "matched_kpi_accounts": 0,
        "unmatched_kpi_ids": [],
        "warnings": [],
        "errors": [str(error)],
        "publishable": False,
        "quality_status": "failed",
    }


def is_quality_publishable(report: dict) -> bool:
    status = report.get("quality_status")
    return status in {None, "ready", "ready_with_warnings"} and not report.get("errors")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build monthly Xiaohongshu insight artifacts.")
    parser.add_argument("--month", required=True, help="Source month in YYYY-MM format.")
    parser.add_argument("--data-root", type=Path, default=_default_data_root())
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parents[1] / "generated")
    parser.add_argument("--quality-only", action="store_true")
    parser.add_argument("--ai", action="store_true", help="Generate validated AI insights or use a validated fallback.")
    parser.add_argument("--no-compat", action="store_true", help="Do not update legacy latest-path JSON files.")
    args = parser.parse_args(argv)
    output_path = args.output_dir / "quality_report.json"
    try:
        month = validate_month(args.month)
        sources = discover_month_sources(args.data_root, month)
        monthly_accounts = [
            (
                item["month"],
                read_monthly_accounts(item["path"], MONTHLY_ACCOUNT_SHEET, item["month"]),
            )
            for item in sources["q4_actual_workbooks"]
        ]
        accounts = next(records for source_month, records in monthly_accounts if source_month == month)
        q4_actuals = aggregate_q4_actuals(monthly_accounts)
        notes = read_notes(sources["monthly_workbook"], MONTHLY_NOTE_SHEET, month)
        kpis = read_kpi(sources["kpi_workbook"], KPI_QUARTER)
        report = profile_quality(accounts, notes, kpis, month, sources)
    except (SourceDiscoveryError, ValueError) as error:
        report = _failed_quality_report(args.month, error)
        _write_json(output_path, report)
        print(f"Quality report failed: {output_path}: {error}", file=sys.stderr)
        return 1

    _write_json(output_path, report)
    print(
        f"Quality report {report['quality_status']}: {output_path} "
        f"({report['account_rows']} accounts, {report['note_rows']} notes)"
    )
    if not is_quality_publishable(report):
        return 1
    if not args.quality_only:
        payload = build_insight_payload(
            accounts, notes, kpis, _category_mapping(), _region_overrides(), report, month, q4_actuals
        )
        attach_ai_artifacts(payload, args.output_dir, use_provider=args.ai)
        if not args.no_compat:
            insight_path = args.output_dir / "insight_data.json"
            _write_json(insight_path, payload)
            write_dealer_scoped_artifacts(args.output_dir, payload)
        write_versioned_artifacts(args.output_dir, payload)
        print(f"Insight data: {args.output_dir / 'months' / month} ({len(payload['dealers'])} dealers)")
    return 0


CITY_REGION_REFERENCE = {
    "北京": ("北京市", "华北"),
    "上海": ("上海市", "华东"),
    "深圳": ("广东省", "华南"),
    "广州": ("广东省", "华南"),
    "成都": ("四川省", "西南"),
    "杭州": ("浙江省", "华东"),
    "南京": ("江苏省", "华东"),
    "苏州": ("江苏省", "华东"),
    "武汉": ("湖北省", "华中"),
    "西安": ("陕西省", "西北"),
    "长春": ("吉林省", "东北"),
    "鄂尔多斯": ("内蒙古自治区", "华北"),
    "天津": ("天津市", "华北"),
    "重庆": ("重庆市", "西南"),
    "郑州": ("河南省", "华中"),
    "济南": ("山东省", "华东"),
    "青岛": ("山东省", "华东"),
    "沈阳": ("辽宁省", "东北"),
    "大连": ("辽宁省", "东北"),
    "哈尔滨": ("黑龙江省", "东北"),
    "昆明": ("云南省", "西南"),
    "贵阳": ("贵州省", "西南"),
    "南宁": ("广西壮族自治区", "华南"),
    "福州": ("福建省", "华南"),
    "厦门": ("福建省", "华南"),
    "合肥": ("安徽省", "华中"),
    "长沙": ("湖南省", "华中"),
    "南昌": ("江西省", "华中"),
    "石家庄": ("河北省", "华北"),
    "太原": ("山西省", "华北"),
    "兰州": ("甘肃省", "西北"),
    "乌鲁木齐": ("新疆维吾尔自治区", "西北"),
    "呼和浩特": ("内蒙古自治区", "华北"),
    "宁波": ("浙江省", "华东"),
    "温州": ("浙江省", "华东"),
    "无锡": ("江苏省", "华东"),
    "佛山": ("广东省", "华南"),
    "东莞": ("广东省", "华南"),
    "珠海": ("广东省", "华南"),
    "海口": ("海南省", "华南"),
    "三亚": ("海南省", "华南"),
}


def normalize_text(value: Any) -> str:
    return " ".join(str(value).split()) if value is not None else ""


def map_category(raw_category: Any, mapping: dict) -> tuple[str, bool]:
    category = normalize_text(raw_category)
    if not category:
        return "未分类", False
    configured = mapping.get(category)
    if configured is None:
        return "其他", False
    return configured["unified"], configured["confirmed"]


def resolve_region(
    dealer: Any,
    store: Any,
    account: Any,
    overrides: dict,
    author_id: Any = "",
) -> dict:
    normalized_overrides = {
        normalize_text(key): value for key, value in overrides.items()
    }
    for value in (account, author_id, dealer, store):
        override = normalized_overrides.get(normalize_text(value))
        if override is not None:
            return {
                "city": override["city"],
                "province": override["province"],
                "region": override["region"],
                "confidence": "confirmed",
            }

    for value in (account, store, dealer):
        search_text = normalize_text(value)
        matches = [city for city in CITY_REGION_REFERENCE if city in search_text]
        if matches:
            city = max(matches, key=len)
            province, region = CITY_REGION_REFERENCE[city]
            return {
                "city": city,
                "province": province,
                "region": region,
                "confidence": "inferred",
            }
    return {"city": "待补充区域", "province": "", "region": "", "confidence": "unknown"}


def classify_account_cohort(author_id: Any, kpi_ids: set) -> str:
    normalized_kpi_ids = {normalize_text(kpi_id) for kpi_id in kpi_ids}
    return "core_kpi" if normalize_text(author_id) in normalized_kpi_ids else "expanded_store"


def safe_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        number = float(value)
        return number if math.isfinite(number) else 0.0
    except (TypeError, ValueError):
        return 0.0


def calculate_kpi_status(actual: float, target: float, elapsed_ratio: float) -> dict:
    actual_value = safe_number(actual)
    target_value = safe_number(target)
    if target_value <= 0:
        return {
            "actual": actual_value,
            "target": target_value,
            "completion_rate": None,
            "pacing_gap": None,
            "status": "unmatched",
        }
    completion = actual_value / target_value
    gap = completion - elapsed_ratio
    if gap >= 0.10:
        status = "leading"
    elif gap >= -0.05:
        status = "normal"
    elif gap >= -0.15:
        status = "warning"
    else:
        status = "critical"
    return {
        "actual": actual_value,
        "target": target_value,
        "completion_rate": completion,
        "pacing_gap": gap,
        "status": status,
    }


def calculate_content_metrics(values: dict) -> dict:
    reads = safe_number(values.get("reads"))
    notes = safe_number(values.get("notes"))
    interactions = sum(
        safe_number(values.get(key)) for key in ("likes", "collects", "comments")
    )
    return {
        "reads": reads,
        "notes": notes,
        "interactions": interactions,
        "reads_per_note": reads / notes if notes else None,
        "interaction_rate": interactions / reads if reads else None,
        "fans_per_10k_reads": (
            safe_number(values.get("new_fans")) / reads * 10000 if reads else None
        ),
        "visitor_rate": safe_number(values.get("visitors")) / reads if reads else None,
        "shares": safe_number(values.get("shares")),
    }


def build_dealer_insights(
    accounts: list[dict],
    notes: list[dict],
    kpis: list[dict],
    category_mapping: dict,
    region_overrides: dict,
    source_month: str,
    q4_actuals: dict[str, dict] | None = None,
) -> list[dict]:
    month = validate_month(source_month)
    kpi_by_author = {
        normalize_text(record.get("author_id")): record
        for record in kpis
        if normalize_text(record.get("author_id"))
    }
    notes_by_author: dict[str, list[dict]] = defaultdict(list)
    for note in notes:
        if _published_in_month(note, month):
            notes_by_author[normalize_text(note.get("author_id"))].append(note)

    grouped_accounts: dict[str, list[dict]] = defaultdict(list)
    for account in accounts:
        name = normalize_text(account.get("dealer")) or normalize_text(account.get("account_name"))
        grouped_accounts[name or "unassigned"].append(account)

    dealers = []
    for dealer_name in sorted(grouped_accounts, key=lambda value: (normalize_text(value), value)):
        child_accounts = sorted(grouped_accounts[dealer_name], key=lambda item: normalize_text(item.get("author_id")))
        account_records = []
        matched_kpis = []
        expanded_values = _empty_values()
        content_notes = []
        cohort_notes: dict[str, list[dict]] = defaultdict(list)
        regional_notes: dict[str, list[dict]] = defaultdict(list)
        regional_cohort_notes: dict[tuple[str, str], list[dict]] = defaultdict(list)
        city_cohort_notes: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
        for account in child_accounts:
            author_id = normalize_text(account.get("author_id"))
            geography = resolve_region(
                account.get("dealer"), account.get("store"), account.get("account_name"), region_overrides, author_id
            )
            cohort = classify_account_cohort(author_id, set(kpi_by_author))
            metrics = _account_metrics(account)
            kpi_record = kpi_by_author.get(author_id)
            account_records.append({
                "author_id": author_id,
                "xhs_id": normalize_text(account.get("xhs_id")),
                "account_name": normalize_text(account.get("account_name")),
                "store": normalize_text(account.get("store")),
                "cohort": cohort,
                **geography,
                "metrics": metrics,
            })
            if kpi_record:
                matched_kpis.append((account, kpi_record, (q4_actuals or {}).get(author_id, metrics)))
            else:
                _add_values(expanded_values, metrics)
            account_notes = notes_by_author.get(author_id, [])
            content_notes.extend(account_notes)
            cohort_notes[cohort].extend(account_notes)
            regional_notes[geography["region"] or "unassigned"].extend(account_notes)
            regional_cohort_notes[(geography["region"] or "unassigned", cohort)].extend(account_notes)
            if geography["confidence"] != "unknown":
                city_cohort_notes[(geography["city"], geography["region"], cohort)].extend(account_notes)

        account_records.sort(key=lambda item: item["author_id"])
        content = _build_content_summary(content_notes, category_mapping)
        content["by_region"] = [
            {"region": region, "categories": _build_content_summary(regional_notes[region], category_mapping)["categories"]}
            for region in sorted(regional_notes)
        ]
        content["by_region_cohort"] = [
            {
                "region": region,
                "cohort": cohort,
                "categories": _build_content_summary(regional_cohort_notes[(region, cohort)], category_mapping)["categories"],
            }
            for region, cohort in sorted(regional_cohort_notes)
        ]
        content["by_city_cohort"] = [
            {
                "city": city,
                "region": region,
                "cohort": cohort,
                "content": _build_content_summary(city_cohort_notes[(city, region, cohort)], category_mapping),
            }
            for city, region, cohort in sorted(city_cohort_notes)
        ]
        kpi_metrics, account_statuses = _build_kpi_metrics(matched_kpis, month)
        dealer = {
            "dealer_id": _stable_id("dealer", dealer_name),
            "name": dealer_name,
            "cohort": "core_kpi" if matched_kpis else "expanded_store",
            "accounts": account_records,
            "kpi": {
                **kpi_metrics,
                "elapsed_ratio": q4_elapsed_ratio(month),
                "account_statuses": account_statuses,
                "overall_status": _overall_status(account_statuses),
            },
            "expanded_store_metrics": calculate_content_metrics(expanded_values),
            "content": content,
            "content_by_cohort": {
                cohort: _build_content_summary(cohort_notes[cohort], category_mapping)
                for cohort in sorted(cohort_notes)
            },
            "recommendations": [],
        }
        dealers.append(dealer)

    _attach_category_benchmarks(dealers)
    for dealer in dealers:
        recommendations = generate_category_recommendations(dealer["dealer_id"], dealer["content"]["categories"])
        recommendations.extend(_kpi_recommendations(dealer))
        recommendations.extend(_dealer_quality_recommendations(dealer))
        dealer["recommendations"] = sorted(recommendations, key=lambda item: item["id"])
    return sorted(dealers, key=lambda item: (item["name"], item["dealer_id"]))


def build_apple_insights(dealers: list[dict], quality_report: dict, source_month: str) -> dict:
    month = validate_month(source_month)
    elapsed = q4_elapsed_ratio(month)
    network_kpis = {}
    for dimension in ("reads", "interactions", "fans"):
        actual = sum(safe_number(dealer.get("kpi", {}).get(dimension, {}).get("actual")) for dealer in dealers)
        target = sum(safe_number(dealer.get("kpi", {}).get(dimension, {}).get("target")) for dealer in dealers)
        network_kpis[dimension] = calculate_kpi_status(actual, target, elapsed)

    statuses = Counter()
    core_accounts = 0
    expanded_accounts = 0
    account_locations = []
    for dealer in dealers:
        account_statuses = dealer.get("kpi", {}).get("account_statuses", [])
        if account_statuses:
            statuses.update(item["status"] for item in account_statuses)
        elif dealer.get("cohort") == "core_kpi":
            statuses[dealer.get("kpi", {}).get("overall_status", "unmatched")] += 1
        for account in dealer.get("accounts", []):
            if account.get("cohort") == "core_kpi":
                core_accounts += 1
            else:
                expanded_accounts += 1
            account_locations.append(account)
    status_counts = {
        key: statuses[key]
        for key in ("leading", "normal", "warning", "critical", "unmatched")
    }
    categories = _aggregate_apple_categories(dealers)
    actions = _synthesize_apple_actions(dealers, _global_quality_recommendations(quality_report))
    return {
        "source_month": month,
        "network_kpis": network_kpis,
        "status_counts": status_counts,
        "account_counts": {"core_kpi": core_accounts, "expanded_store": expanded_accounts},
        "dealer_quadrants": _dealer_quadrants(dealers),
        "regional_summaries": _location_summaries(account_locations, "region"),
        "city_summaries": _location_summaries(account_locations, "city"),
        "category_mix_performance": categories,
        "risk_dealers": _risk_dealers(dealers),
        "replicable_cases": _replicable_cases(dealers, categories),
        "actions": actions,
        "quality_metadata": {
            "matched_kpi_accounts": safe_number(quality_report.get("matched_kpi_accounts")),
            "unmatched_kpi_accounts": len(quality_report.get("unmatched_kpi_ids", [])),
            "data_freshness": quality_report.get("data_freshness"),
        },
    }


def _risk_dealers(dealers: list[dict]) -> list[dict]:
    severity_rank = {"critical": 0, "warning": 1}
    risks = []
    for dealer in dealers:
        status = dealer.get("kpi", {}).get("overall_status")
        if status not in severity_rank:
            continue
        gaps = [
            metric.get("pacing_gap")
            for dimension in ("reads", "interactions", "fans")
            if isinstance((metric := dealer.get("kpi", {}).get(dimension, {})).get("pacing_gap"), (int, float))
            and math.isfinite(metric["pacing_gap"])
        ]
        risks.append({
            "dealer_id": dealer["dealer_id"],
            "name": dealer["name"],
            "status": status,
            "worst_pacing_gap": min(gaps) if gaps else None,
            "risk_basis": "minimum_q4_kpi_pacing_gap",
        })
    return sorted(
        risks,
        key=lambda item: (
            severity_rank[item["status"]],
            item["worst_pacing_gap"] if item["worst_pacing_gap"] is not None else math.inf,
            normalize_text(item["name"]),
            item["dealer_id"],
        ),
    )


def build_insight_payload(
    accounts: list[dict],
    notes: list[dict],
    kpis: list[dict],
    category_mapping: dict,
    region_overrides: dict,
    quality_report: dict,
    source_month: str,
    q4_actuals: dict[str, dict] | None = None,
) -> dict:
    month = validate_month(source_month)
    if not is_quality_publishable(quality_report):
        raise ValueError("Quality report is not publishable")
    dealers = build_dealer_insights(accounts, notes, kpis, category_mapping, region_overrides, month, q4_actuals)
    payload_quality = {key: value for key, value in quality_report.items() if key != "generated_at"}
    payload_quality.setdefault("quality_status", "ready")
    payload_quality.setdefault("errors", [])
    payload_quality["publishable"] = True
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    payload = {
        "schema_version": "2.0",
        "source_month": month,
        "previous_month": previous_month(month),
        "generated_at": generated_at,
        "metadata": {
            "generated_at": generated_at,
            "data_freshness": quality_report.get("data_freshness"),
            "logic_version": LOGIC_VERSION,
        },
        "source_files": quality_report.get("source_files", {}),
        "quality": payload_quality,
        "analysis_rules": {
            "interactions": "likes_plus_collects_plus_comments_excluding_shares",
            "content_scope": "notes_published_in_source_month_current_snapshot_metrics",
            "benchmark": "same_cohort_median_with_low_confidence_fallback",
            "q4_elapsed_ratio": q4_elapsed_ratio(month),
            "quarter_forecast": "not_shown_until_two_completed_q4_months",
        },
        "dealers": dealers,
        "apple": build_apple_insights(dealers, quality_report, month),
    }
    payload = _clean_non_finite(payload)
    json.dumps(payload, ensure_ascii=False, allow_nan=False)
    return payload


def summarize_payload_for_history(payload: dict) -> dict:
    apple = payload.get("apple") or {}
    network = apple.get("network_kpis") or {}
    reads = network.get("reads") if isinstance(network.get("reads"), dict) else {}
    interactions = network.get("interactions") if isinstance(network.get("interactions"), dict) else {}
    fans = network.get("fans") if isinstance(network.get("fans"), dict) else {}
    dealer_contents = [
        dealer.get("content") for dealer in payload.get("dealers", [])
        if isinstance(dealer, dict) and isinstance(dealer.get("content"), dict)
    ]
    notes = sum(safe_number(content.get("notes")) for content in dealer_contents)
    content_reads = sum(safe_number(content.get("reads")) for content in dealer_contents)
    content_interactions = sum(safe_number(content.get("interactions")) for content in dealer_contents)
    content_fans = sum(safe_number(content.get("new_fans")) for content in dealer_contents)
    categories = apple.get("category_mix_performance") or []
    category_reads_per_note = [
        safe_number(item.get("reads_per_note")) for item in categories
        if isinstance(item, dict) and item.get("reads_per_note") is not None
    ]
    category_baseline = median(category_reads_per_note) if category_reads_per_note else None
    viral_categories = [
        item for item in categories
        if isinstance(item, dict) and category_baseline is not None and safe_number(item.get("reads_per_note")) >= category_baseline
    ]
    return {
        "month": payload["source_month"],
        "reads": safe_number(reads.get("actual") if reads else content_reads),
        "notes": notes,
        "interactions": safe_number(interactions.get("actual") if interactions else content_interactions),
        "new_fans": safe_number(fans.get("actual") if fans else content_fans),
        "viral_rate": len(viral_categories) / len(categories) if categories else None,
    }


def load_history_rows(output_dir: Path, current_month: str) -> list[dict]:
    index = load_month_index(Path(output_dir) / "month_index.json")
    rows = []
    for month in sorted(month for month in index.get("months", []) if isinstance(month, str) and month < current_month):
        apple_path = Path(output_dir) / "months" / month / "apple.json"
        try:
            historical_payload = json.loads(apple_path.read_text(encoding="utf-8"))
        except (FileNotFoundError, json.JSONDecodeError):
            continue
        if isinstance(historical_payload, dict):
            try:
                rows.append(summarize_payload_for_history(historical_payload))
            except (KeyError, TypeError, ValueError):
                continue
    return rows


def attach_ai_artifacts(payload: dict, output_dir: Path, *, use_provider: bool) -> None:
    prompt = load_prompt()
    period = apple_period(payload["source_month"])
    history = build_history_context(summarize_payload_for_history(payload), load_history_rows(output_dir, payload["source_month"]))
    payload["period"] = period
    payload["history"] = history
    payload["metadata"]["logic_version"] = LOGIC_VERSION
    payload["metadata"]["prompt_version"] = prompt["prompt_version"]

    env = os.environ if use_provider else {}
    cache_root = Path(output_dir) / ".ai_cache" / payload["source_month"]
    apple_packet = build_evidence_packet("apple", payload, history, period)
    payload["apple"]["evidence"] = apple_packet["evidence"]
    payload["apple"]["ai_insights"] = generate_ai_insights(
        apple_packet,
        cache_path=cache_root / "apple.json",
        env=env,
    )
    for dealer in payload.get("dealers", []):
        scoped_payload = build_dealer_payload(payload, dealer["dealer_id"])
        dealer_packet = build_evidence_packet("dealer", scoped_payload, history, period)
        dealer["evidence"] = dealer_packet["evidence"]
        dealer["ai_insights"] = generate_ai_insights(
            dealer_packet,
            cache_path=cache_root / "dealers" / f"{dealer['dealer_id']}.json",
            env=env,
        )


def q4_elapsed_ratio(month: str) -> float:
    current = datetime.strptime(validate_month(month), "%Y-%m")
    if current.month in {7, 8, 9}:
        return (current.month - 6) / 3
    return 1.0


def generate_category_recommendations(dealer_id: str, categories: list[dict]) -> list[dict]:
    recommendations = []
    for category in sorted(categories, key=lambda item: item["category"]):
        name = category["category"]
        notes = safe_number(category.get("notes"))
        supply = safe_number(category.get("note_share"))
        efficiency = safe_number(category.get("reads_per_note"))
        supply_benchmark = safe_number(category.get("benchmark_note_share"))
        efficiency_benchmark = safe_number(category.get("benchmark_reads_per_note"))
        evidence = [
            {"metric": "note_share", "value": supply, "benchmark": supply_benchmark, "scope": "same_cohort"},
            {"metric": "reads_per_note", "value": efficiency, "benchmark": efficiency_benchmark, "scope": "same_cohort"},
            {"metric": "notes", "value": notes, "benchmark": 3, "scope": "dealer_category"},
        ]
        target = {"category": name, "city": "", "account_id": ""}
        if safe_number(category.get("mapping_completeness", 1)) < 1:
            recommendations.append(_recommendation(
                "data_quality", dealer_id, name, "\u786e\u8ba4\u5185\u5bb9\u5206\u7c7b\u6620\u5c04",
                "\u5148\u786e\u8ba4\u539f\u59cb\u5206\u7c7b\u6620\u5c04\uff0c\u4e0d\u5bf9\u672a\u786e\u8ba4\u5206\u7c7b\u7ed9\u51fa\u5185\u5bb9\u7b56\u7565\u3002", "validate", "low",
                evidence + [{"metric": "mapping_completeness", "value": category.get("mapping_completeness"), "benchmark": 1, "scope": "dealer_category"}],
                {"category": "", "city": "", "account_id": ""}, "quality_category_mapping",
            ))
            continue
        if notes < 3 or category.get("benchmark_confidence") != "supported":
            title, action, confidence, priority, rule_id = (
                f"\u9a8c\u8bc1{name}\u5185\u5bb9\u6837\u672c",
                f"\u4e0b\u6708\u8865\u9f50{name}\u81f3\u5c113\u7bc7\u540e\u590d\u6838\u4f9b\u7ed9\u4e0e\u9605\u8bfb\u6548\u7387\u3002",
                "validate", "low", "category_validate",
            )
        elif supply < supply_benchmark and efficiency > efficiency_benchmark:
            title, action, confidence, priority, rule_id = (
                f"\u503c\u5f97\u52a0\u7801{name}",
                f"\u4e0b\u6708\u589e\u52a0{name}\u4f9b\u7ed9\u81f3\u540c\u961f\u5217\u4e2d\u4f4d\u6c34\u5e73\u3002",
                "supported", "medium", "category_scale",
            )
        elif supply >= supply_benchmark and efficiency < efficiency_benchmark:
            title, action, confidence, priority, rule_id = (
                f"\u4f18\u5316{name}\u5185\u5bb9",
                f"\u4fdd\u6301{name}\u4f9b\u7ed9\uff0c\u6d4b\u8bd5\u9009\u9898\u4e0e\u5c01\u9762\u4ee5\u63d0\u5347\u5355\u7bc7\u9605\u8bfb\u3002",
                "supported", "medium", "category_optimize",
            )
        elif supply >= supply_benchmark and efficiency >= efficiency_benchmark:
            title, action, confidence, priority, rule_id = (
                f"\u7a33\u5b9a\u4fdd\u6301{name}",
                f"\u7ee7\u7eed\u4fdd\u6301{name}\u4f9b\u7ed9\u8282\u594f\uff0c\u5e76\u8bb0\u5f55\u53ef\u590d\u7528\u7d20\u6750\u3002",
                "signal", "low", "category_maintain",
            )
        else:
            title, action, confidence, priority, rule_id = (
                f"\u9a8c\u8bc1{name}\u5185\u5bb9\u6548\u7387",
                f"\u4e0b\u6708\u4fdd\u6301{name}\u5c0f\u8303\u56f4\u6d4b\u8bd5\uff0c\u786e\u8ba4\u4f9b\u7ed9\u4e0e\u6548\u7387\u5173\u7cfb\u3002",
                "validate", "low", "category_validate",
            )
        recommendations.append(_recommendation(
            "category", dealer_id, name, title, action, confidence, priority, evidence, target, rule_id
        ))
    return recommendations


def _stable_id(prefix: str, value: str) -> str:
    digest = hashlib.sha1(normalize_text(value).encode("utf-8")).hexdigest()[:12]
    return f"{prefix}-{digest}"


def _empty_values() -> dict:
    return {key: 0.0 for key in ("reads", "likes", "collects", "comments", "shares", "new_fans", "visitors", "notes")}


def _add_values(target: dict, source: dict) -> None:
    for key in target:
        target[key] += safe_number(source.get(key))


def _account_metrics(account: dict) -> dict:
    return {
        key: safe_number(account.get(key))
        for key in ("reads", "likes", "collects", "comments", "shares", "new_fans", "visitors")
    }


def _published_in_month(note: dict, source_month: str) -> bool:
    published = note.get("publish_date", note.get("published_at", ""))
    if isinstance(published, datetime):
        return published.strftime("%Y-%m") == source_month
    return normalize_text(published)[:7] == source_month


def _build_content_summary(notes: list[dict], category_mapping: dict) -> dict:
    values = _empty_values()
    category_values: dict[str, dict] = {}
    raw_categories: Counter = Counter()
    image_notes = 0
    video_notes = 0
    seen_note_ids = set()
    for note in sorted(notes, key=lambda item: normalize_text(item.get("note_id"))):
        note_id = normalize_text(note.get("note_id"))
        if note_id and note_id in seen_note_ids:
            continue
        seen_note_ids.add(note_id)
        note_values = _account_metrics(note)
        note_values["notes"] = 1.0
        _add_values(values, note_values)
        raw_category = normalize_text(note.get("category"))
        unified, confirmed = map_category(raw_category, category_mapping)
        raw_categories[raw_category or "(blank)"] += 1
        item = category_values.setdefault(unified, {
            "category": unified,
            "raw_categories": Counter(),
            "mapping_confirmed_notes": 0,
            "blank_category_notes": 0,
            "values": _empty_values(),
        })
        item["raw_categories"][raw_category or "(blank)"] += 1
        item["mapping_confirmed_notes"] += int(confirmed)
        item["blank_category_notes"] += int(not raw_category)
        _add_values(item["values"], note_values)
        format_value = normalize_text(note.get("note_format")).lower()
        if "\u89c6\u9891" in format_value or format_value == "video":
            video_notes += 1
        else:
            image_notes += 1

    content = calculate_content_metrics(values)
    content["new_fans"] = values["new_fans"]
    content["methodology_note"] = "\u5f53\u6708\u65b0\u53d1\u7b14\u8bb0\u4f7f\u7528\u5f53\u524d\u7d2f\u8ba1\u5feb\u7167\u6307\u6807\uff0c\u4e0d\u89e3\u8bfb\u4e3a\u5f53\u6708\u56e0\u679c\u589e\u91cf\u3002"
    content["image_notes"] = image_notes
    content["video_notes"] = video_notes
    content["image_share"] = image_notes / values["notes"] if values["notes"] else None
    content["video_share"] = video_notes / values["notes"] if values["notes"] else None
    content["raw_category_counts"] = dict(sorted(raw_categories.items()))
    categories = []
    for name in sorted(category_values):
        item = category_values[name]
        metrics = calculate_content_metrics(item["values"])
        note_count = item["values"]["notes"]
        categories.append({
            "category": name,
            "raw_categories": dict(sorted(item["raw_categories"].items())),
            "notes": note_count,
            "note_share": note_count / values["notes"] if values["notes"] else None,
            "reads": metrics["reads"],
            "interactions": metrics["interactions"],
            "shares": metrics["shares"],
            "new_fans": item["values"]["new_fans"],
            "reads_per_note": metrics["reads_per_note"],
            "interaction_rate": metrics["interaction_rate"],
            "fans_per_10k_reads": metrics["fans_per_10k_reads"],
            "mapping_completeness": item["mapping_confirmed_notes"] / note_count if note_count else None,
            "blank_category_notes": item["blank_category_notes"],
        })
    content["categories"] = categories
    return content


def _build_kpi_metrics(matched_kpis: list[tuple[dict, dict, dict]], source_month: str) -> tuple[dict, list[dict]]:
    elapsed = q4_elapsed_ratio(source_month)
    metric_records = {dimension: {"actual": 0.0, "target": 0.0} for dimension in ("reads", "interactions", "fans")}
    account_statuses = []
    for account, kpi, metrics in matched_kpis:
        account_dimensions = {
            "reads": (metrics["reads"], kpi.get("read_target")),
            "interactions": (metrics["likes"] + metrics["collects"] + metrics["comments"], kpi.get("interaction_target")),
            "fans": (metrics["new_fans"], kpi.get("fan_target")),
        }
        statuses = []
        for dimension, (actual, target) in account_dimensions.items():
            metric_records[dimension]["actual"] += safe_number(actual)
            metric_records[dimension]["target"] += safe_number(target)
            statuses.append(calculate_kpi_status(actual, target, elapsed)["status"])
        account_statuses.append({
            "author_id": normalize_text(account.get("author_id")),
            "status": _worst_status(statuses),
            "actual_source_months": list(metrics.get("source_months", [source_month])),
        })
    return (
        {dimension: calculate_kpi_status(values["actual"], values["target"], elapsed) for dimension, values in metric_records.items()},
        sorted(account_statuses, key=lambda item: item["author_id"]),
    )


def _worst_status(statuses: list[str]) -> str:
    ranks = {"leading": 0, "normal": 1, "warning": 2, "critical": 3, "unmatched": 4}
    return max(statuses, key=lambda item: ranks[item]) if statuses else "unmatched"


def _overall_status(account_statuses: list[dict]) -> str:
    return _worst_status([item["status"] for item in account_statuses])


def _attach_category_benchmarks(dealers: list[dict]) -> None:
    for dealer in dealers:
        cohort_dealers = [item for item in dealers if item["cohort"] == dealer["cohort"]]
        category_by_dealer = {
            item["dealer_id"]: {category["category"]: category for category in item["content"]["categories"]}
            for item in cohort_dealers
        }
        for category in dealer["content"]["categories"]:
            comparable = [
                categories[category["category"]]
                for categories in category_by_dealer.values()
                if category["category"] in categories
            ]
            if len(comparable) >= 3:
                category["benchmark_note_share"] = median(safe_number(item.get("note_share")) for item in comparable)
                category["benchmark_reads_per_note"] = median(safe_number(item.get("reads_per_note")) for item in comparable)
                category["benchmark_interaction_rate"] = median(safe_number(item.get("interaction_rate")) for item in comparable)
                confidence = "supported"
            else:
                category["benchmark_note_share"] = median(
                    safe_number(category_by_dealer[item["dealer_id"]].get(category["category"], {}).get("note_share"))
                    for item in cohort_dealers
                )
                category["benchmark_reads_per_note"] = median(
                    safe_number(item["content"].get("reads_per_note")) for item in cohort_dealers
                )
                category["benchmark_interaction_rate"] = median(
                    safe_number(item["content"].get("interaction_rate")) for item in cohort_dealers
                )
                confidence = "lower_confidence"
            category["benchmark_confidence"] = confidence
            category["benchmark_sample_size"] = len(comparable)


def _recommendation(
    recommendation_type: str, dealer_id: str, subject: str, title: str, action: str,
    confidence: str, priority: str, evidence: list[dict], target: dict, rule_id: str | None = None,
) -> dict:
    return {
        "id": _stable_id(recommendation_type, f"{dealer_id}:{rule_id or recommendation_type}:{subject}:{title}"),
        "rule_id": rule_id or recommendation_type,
        "type": recommendation_type,
        "title": title,
        "action": action,
        "confidence": confidence,
        "priority": priority,
        "evidence": evidence,
        "target": target,
    }


def _kpi_recommendations(dealer: dict) -> list[dict]:
    recommendations = []
    for dimension in ("reads", "interactions", "fans"):
        values = dealer["kpi"][dimension]
        if values["status"] not in {"warning", "critical"}:
            continue
        evidence = [
            {"metric": f"{dimension}_actual", "value": values["actual"], "benchmark": values["target"], "scope": "dealer_core_kpi"},
            {"metric": f"{dimension}_completion", "value": values["completion_rate"], "benchmark": dealer["kpi"]["elapsed_ratio"], "scope": "q4_time_progress"},
            {"metric": f"{dimension}_pacing_gap", "value": values["pacing_gap"], "benchmark": 0, "scope": "q4_time_progress"},
        ]
        recommendations.append(_recommendation(
            "kpi", dealer["dealer_id"], f"{dimension}:{values['status']}",
            f"\u63a8\u8fdb{dealer['name']}{dimension}KPI\u8282\u594f", "\u7acb\u5373\u6839\u636e\u65f6\u95f4\u8fdb\u5ea6\u62c6\u89e3\u672c\u6708\u52a8\u4f5c\u3002",
            "supported", "high", evidence, {"category": "", "city": "", "account_id": ""}, f"kpi_{dimension}_{values['status']}",
        ))
    return recommendations


def _dealer_quality_recommendations(dealer: dict) -> list[dict]:
    recommendations = []
    for account in dealer["accounts"]:
        if account["confidence"] == "unknown":
            recommendations.append(_recommendation(
                "data_quality", dealer["dealer_id"], account["author_id"], "\u8865\u5168\u8d26\u53f7\u57ce\u5e02\u4fe1\u606f",
                "\u5148\u6838\u5bf9\u8d26\u53f7\u6240\u5c5e\u57ce\u5e02\uff0c\u518d\u5f00\u5c55\u533a\u57df\u6216\u57ce\u5e02\u7b56\u7565\u5206\u6790\u3002", "validate", "low",
                [{"metric": "city_confidence", "value": 0, "benchmark": 1, "scope": "account"}],
                {"category": "", "city": "", "account_id": account["author_id"]}, "quality_unknown_city",
            ))
    return recommendations


def _priority_rank(priority: str) -> int:
    return {"high": 3, "medium": 2, "low": 1}.get(priority, 0)


def _action_context(dealer: dict | None, recommendation: dict) -> tuple[str, str]:
    if dealer is None:
        return "network", "network"
    target_account = recommendation.get("target", {}).get("account_id")
    matching_account = next(
        (account for account in dealer.get("accounts", []) if account.get("author_id") == target_account),
        None,
    )
    if matching_account is not None:
        return matching_account.get("region") or "unassigned", matching_account.get("cohort", dealer["cohort"])
    regions = sorted({account.get("region") or "unassigned" for account in dealer.get("accounts", [])})
    return (regions[0] if len(regions) == 1 else "multi_region"), dealer["cohort"]


def _aggregate_action_evidence(recommendations: list[dict], affected_dealers: int) -> list[dict]:
    evidence = [
        {"metric": "affected_dealer_count", "value": affected_dealers, "benchmark": 0, "scope": "network_action"},
        {"metric": "recommendation_count", "value": len(recommendations), "benchmark": 0, "scope": "network_action"},
    ]
    metrics: dict[str, list[dict]] = defaultdict(list)
    for recommendation in recommendations:
        for item in recommendation.get("evidence", []):
            metrics[item.get("metric", "unknown")].append(item)
    for metric in sorted(metrics)[:3]:
        items = metrics[metric]
        values = [safe_number(item.get("value")) for item in items]
        benchmarks = [safe_number(item.get("benchmark")) for item in items]
        evidence.append({
            "metric": metric,
            "value": sum(values) / len(values),
            "benchmark": sum(benchmarks) / len(benchmarks),
            "scope": "aggregated_recommendations",
        })
    return evidence


def _apple_action_copy(rule_id: str, category: str, region: str, cohort: str) -> tuple[str, str]:
    scope = f"{region}{cohort}" if region not in {"network", "multi_region"} else "\u7f51\u7edc"
    if rule_id == "category_scale":
        return f"\u7f51\u7edc\u52a0\u7801{category}\u5185\u5bb9", f"\u5728{scope}\u6269\u5927{category}\u4f9b\u7ed9\uff0c\u5e76\u4fdd\u7559\u53ef\u590d\u7528\u7d20\u6750\u3002"
    if rule_id == "category_optimize":
        return f"\u7f51\u7edc\u4f18\u5316{category}\u5185\u5bb9", f"\u5728{scope}\u4fdd\u6301{category}\u4f9b\u7ed9\uff0c\u96c6\u4e2d\u6d4b\u8bd5\u9009\u9898\u4e0e\u5c01\u9762\u3002"
    if rule_id == "category_maintain":
        return f"\u7f51\u7edc\u4fdd\u6301{category}\u4f18\u52bf", f"\u5728{scope}\u7ee7\u7eed\u4fdd\u6301{category}\u4f9b\u7ed9\u8282\u594f\uff0c\u6c89\u6dc0\u53ef\u590d\u7528\u505a\u6cd5\u3002"
    if rule_id == "category_validate":
        return f"\u7f51\u7edc\u9a8c\u8bc1{category}\u65b9\u5411", f"\u5728{scope}\u8865\u8db3\u6837\u672c\u540e\u9a8c\u8bc1{category}\u4f9b\u7ed9\u4e0e\u6548\u7387\u3002"
    if rule_id.startswith("kpi_"):
        _, metric, status = rule_id.split("_", 2)
        metric_label = {"reads": "\u9605\u8bfb", "interactions": "\u4e92\u52a8", "fans": "\u6da8\u7c89"}.get(metric, metric)
        return f"\u7f51\u7edc\u63a8\u8fdb{metric_label}KPI", f"\u56f4\u7ed5{metric_label}\u6307\u6807\u62c6\u89e3{status}\u8d26\u53f7\u7684\u65f6\u95f4\u8282\u594f\u4e0e\u8f6c\u5316\u52a8\u4f5c\u3002"
    quality_labels = {
        "quality_unknown_city": "\u57ce\u5e02\u4fe1\u606f",
        "quality_unmatched_kpi": "\u672a\u5339\u914dKPI",
        "quality_category_mapping": "\u5206\u7c7b\u6620\u5c04",
    }
    return f"\u7f51\u7edc\u6838\u5bf9{quality_labels.get(rule_id, '\u6570\u636e\u8d28\u91cf')}", f"\u5148\u5b8c\u6210{scope}\u8303\u56f4\u7684\u6570\u636e\u6838\u5bf9\uff0c\u518d\u8fdb\u884c\u7b56\u7565\u5206\u6790\u3002"


def _synthesize_apple_actions(dealers: list[dict], global_recommendations: list[dict], max_per_bucket: int = 12) -> dict:
    grouped: dict[tuple[str, str, str, str, str], list[tuple[dict, dict | None]]] = defaultdict(list)
    for dealer in dealers:
        for recommendation in dealer.get("recommendations", []):
            region, cohort = _action_context(dealer, recommendation)
            category = recommendation.get("target", {}).get("category", "")
            rule_id = recommendation.get("rule_id", recommendation["type"])
            grouped[(rule_id, recommendation["type"], category, region, cohort)].append((recommendation, dealer))
    for recommendation in global_recommendations:
        rule_id = recommendation.get("rule_id", recommendation["type"])
        grouped[(rule_id, recommendation["type"], "", "network", "network")].append((recommendation, None))

    actions = []
    for key in sorted(grouped):
        rule_id, recommendation_type, category, region, cohort = key
        members = sorted(grouped[key], key=lambda item: (-_priority_rank(item[0]["priority"]), item[0]["id"]))
        recommendations = [item[0] for item in members]
        dealer_ids = sorted({dealer["dealer_id"] for _, dealer in members if dealer is not None})
        account_ids = sorted({
            recommendation.get("target", {}).get("account_id", "")
            for recommendation in recommendations
            if recommendation.get("target", {}).get("account_id")
        })
        highest_priority = max(recommendations, key=lambda item: _priority_rank(item["priority"]))["priority"]
        confidence = "validate" if all(item["confidence"] == "validate" for item in recommendations) else (
            "supported" if any(item["confidence"] == "supported" for item in recommendations) else "signal"
        )
        title, action = _apple_action_copy(rule_id, category, region, cohort)
        actions.append({
            "id": _stable_id("apple-action", ":".join(key)),
            "rule_id": rule_id,
            "type": recommendation_type,
            "title": title,
            "action": action,
            "confidence": confidence,
            "priority": highest_priority,
            "region": region,
            "cohort": cohort,
            "affected_dealer_count": len(dealer_ids),
            "affected_account_count": len(account_ids),
            "evidence": _aggregate_action_evidence(recommendations, len(dealer_ids)),
            "target": {"category": category, "city": "", "account_id": ""},
            "top_examples": [
                {
                    "dealer_id": dealer["dealer_id"] if dealer is not None else "",
                    "dealer_name": dealer["name"] if dealer is not None else "network",
                    "recommendation_id": recommendation["id"],
                    "evidence": recommendation.get("evidence", [])[:2],
                }
                for recommendation, dealer in members[:3]
            ],
            "drilldown_recommendation_ids": [item["id"] for item in recommendations],
        })

    ordered = sorted(
        actions,
        key=lambda item: (-_priority_rank(item["priority"]), -item["affected_dealer_count"], -len(item["drilldown_recommendation_ids"]), item["id"]),
    )
    return {
        "\u7acb\u5373\u884c\u52a8": [item for item in ordered if item["priority"] == "high"][:max_per_bucket],
        "\u4e0b\u6708\u9a8c\u8bc1": [item for item in ordered if item["confidence"] == "validate" and item["priority"] != "high"][:max_per_bucket],
        "\u5b63\u5ea6\u89c4\u5212": [item for item in ordered if item["confidence"] != "validate" and item["priority"] != "high"][:max_per_bucket],
    }


def _aggregate_apple_categories(dealers: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str, str], dict] = {}
    for dealer in dealers:
        regional_categories = dealer["content"].get(
            "by_region_cohort",
            [{"region": "network", "cohort": dealer["cohort"], "categories": dealer["content"]["categories"]}],
        )
        for region_summary in regional_categories:
            for category in region_summary["categories"]:
                key = (region_summary["region"], region_summary["cohort"], category["category"])
                item = grouped.setdefault(key, {"region": key[0], "cohort": key[1], "category": key[2], "notes": 0.0, "reads": 0.0, "interactions": 0.0, "new_fans": 0.0})
                for metric in ("notes", "reads", "interactions", "new_fans"):
                    item[metric] += safe_number(category.get(metric))
    result = []
    for item in grouped.values():
        item["reads_per_note"] = item["reads"] / item["notes"] if item["notes"] else None
        item["interaction_rate"] = item["interactions"] / item["reads"] if item["reads"] else None
        item["fans_per_10k_reads"] = item["new_fans"] / item["reads"] * 10000 if item["reads"] else None
        result.append(item)
    return sorted(result, key=lambda item: (item["region"], item["cohort"], item["category"]))


def _dealer_quadrants(dealers: list[dict]) -> list[dict]:
    records = []
    for cohort in sorted({dealer["cohort"] for dealer in dealers}):
        members = [dealer for dealer in dealers if dealer["cohort"] == cohort]
        cohort_content = {
            item["dealer_id"]: item.get("content_by_cohort", {}).get(cohort, item["content"])
            for item in members
        }
        supply_median = median(safe_number(cohort_content[item["dealer_id"]].get("notes")) for item in members) if members else 0
        efficiency_median = median(safe_number(cohort_content[item["dealer_id"]].get("reads_per_note")) for item in members) if members else 0
        for dealer in members:
            content = cohort_content[dealer["dealer_id"]]
            supply = safe_number(content.get("notes"))
            reads = safe_number(content.get("reads"))
            efficiency = safe_number(content.get("reads_per_note"))
            high_supply = supply >= supply_median
            high_efficiency = efficiency >= efficiency_median
            quadrant = (
                "high_supply_high_efficiency" if high_supply and high_efficiency else
                "high_supply_low_efficiency" if high_supply else
                "low_supply_high_efficiency" if high_efficiency else "low_supply_low_efficiency"
            )
            records.append({
                "dealer_id": dealer["dealer_id"], "name": dealer["name"], "cohort": cohort,
                "notes": supply,
                "reads": reads,
                "reads_per_note": efficiency,
                "source_account_ids": sorted(
                    account.get("author_id", "")
                    for account in dealer.get("accounts", [])
                    if account.get("cohort") == cohort and account.get("author_id")
                ),
                "normalized_supply": supply / supply_median if supply_median else None,
                "normalized_efficiency": efficiency / efficiency_median if efficiency_median else None,
                "quadrant": quadrant,
            })
    return sorted(records, key=lambda item: item["dealer_id"])


def _location_summaries(accounts: list[dict], field: str) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for account in accounts:
        grouped[normalize_text(account.get(field)) or "unassigned"].append(account)
    result = []
    for name, members in grouped.items():
        identified = sum(item.get("confidence") != "unknown" for item in members)
        result.append({
            field: name,
            "account_count": len(members),
            "identified_coverage": identified / len(members) if members else None,
        })
    return sorted(result, key=lambda item: item[field])


def _replicable_cases(dealers: list[dict], categories: list[dict]) -> dict:
    cohort_efficiency = {
        cohort: median(safe_number(item["content"].get("reads_per_note")) for item in dealers if item["cohort"] == cohort)
        for cohort in {item["cohort"] for item in dealers}
    }
    successful_dealers = [
        dealer for dealer in dealers
        if safe_number(dealer["content"].get("notes")) >= 3
        and safe_number(dealer["content"].get("reads_per_note")) > cohort_efficiency[dealer["cohort"]]
    ]
    account_cases = []
    for dealer in successful_dealers:
        for account in dealer["accounts"]:
            if account["cohort"] == "core_kpi" and any(
                status.get("author_id") == account.get("author_id") and status.get("status") == "leading"
                for status in dealer["kpi"].get("account_statuses", [])
            ):
                account_cases.append({"dealer_id": dealer["dealer_id"], "account_id": account["author_id"], "evidence": [{"metric": "kpi_status", "value": 1, "benchmark": 0, "scope": "core_kpi"}, {"metric": "reads_per_note", "value": dealer["content"]["reads_per_note"], "benchmark": cohort_efficiency[dealer["cohort"]], "scope": "dealer_content"}]})
    category_cases = []
    for dealer in dealers:
        for category in dealer["content"].get("categories", []):
            if (
                safe_number(category.get("notes")) >= 3
                and category.get("benchmark_confidence") == "supported"
                and safe_number(category.get("mapping_completeness")) == 1
                and safe_number(category.get("reads_per_note")) > safe_number(category.get("benchmark_reads_per_note"))
                and safe_number(category.get("note_share")) >= safe_number(category.get("benchmark_note_share"))
            ):
                category_cases.append({
                    "dealer_id": dealer["dealer_id"], "cohort": dealer["cohort"], "category": category["category"],
                    "evidence": [
                        {"metric": "reads_per_note", "value": category["reads_per_note"], "benchmark": category["benchmark_reads_per_note"], "scope": "dealer_category"},
                        {"metric": "notes", "value": category["notes"], "benchmark": 3, "scope": "dealer_category"},
                    ],
                })
    city_summaries: dict[tuple[str, str, str], dict] = {}
    for dealer in dealers:
        for segment in dealer["content"].get("by_city_cohort", []):
            key = (segment["city"], segment["region"], segment["cohort"])
            item = city_summaries.setdefault(key, {
                "city": key[0], "region": key[1], "cohort": key[2],
                "notes": 0.0, "reads": 0.0, "dealer_ids": set(),
            })
            item["notes"] += safe_number(segment["content"].get("notes"))
            item["reads"] += safe_number(segment["content"].get("reads"))
            item["dealer_ids"].add(dealer["dealer_id"])
    for item in city_summaries.values():
        item["reads_per_note"] = item["reads"] / item["notes"] if item["notes"] else None
    city_benchmarks = {
        cohort: median(safe_number(item.get("reads_per_note")) for item in city_summaries.values() if item["cohort"] == cohort)
        for cohort in {item["cohort"] for item in city_summaries.values()}
    }
    city_cases = [
        {
            "city": item["city"], "region": item["region"], "cohort": item["cohort"],
            "dealer_ids": sorted(item["dealer_ids"]),
            "evidence": [
                {"metric": "reads_per_note", "value": item["reads_per_note"], "benchmark": city_benchmarks[item["cohort"]], "scope": "city_content"},
                {"metric": "notes", "value": item["notes"], "benchmark": 3, "scope": "city_content"},
            ],
        }
        for item in city_summaries.values()
        if safe_number(item["notes"]) >= 3 and safe_number(item.get("reads_per_note")) > city_benchmarks[item["cohort"]]
    ]
    return {
        "category": sorted(category_cases, key=lambda item: (item["cohort"], item["category"], item["dealer_id"])),
        "city": sorted(city_cases, key=lambda item: (item["city"], item["cohort"])),
        "account": sorted(account_cases, key=lambda item: (item["dealer_id"], item["account_id"])),
    }


def _global_quality_recommendations(quality_report: dict) -> list[dict]:
    unmatched = sorted(quality_report.get("unmatched_kpi_ids", []))
    if not unmatched:
        return []
    return [_recommendation(
        "data_quality", "apple", "unmatched_kpi", "\u6838\u5bf9\u672a\u5339\u914dKPI\u8d26\u53f7",
        "\u5148\u6838\u5bf9\u672a\u5339\u914d\u8d26\u53f7\u6807\u8bc6\uff0c\u4e0d\u5c06\u5176\u7f16\u5165\u6838\u5fc3KPI\u6216\u57ce\u5e02\u7b56\u7565\u3002", "validate", "low",
        [{"metric": "unmatched_kpi_accounts", "value": len(unmatched), "benchmark": 0, "scope": "network"}],
        {"category": "", "city": "", "account_id": ""}, "quality_unmatched_kpi",
    )]


def _clean_non_finite(value: Any) -> Any:
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {key: _clean_non_finite(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_clean_non_finite(item) for item in value]
    return value


if __name__ == "__main__":
    raise SystemExit(main())
