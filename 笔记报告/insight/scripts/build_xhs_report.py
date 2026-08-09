from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from .build_insight import (
        MONTHLY_ACCOUNT_SHEET,
        MONTHLY_NOTE_SHEET,
        _discover_monthly_workbook,
        _find_header_row,
        _identifier,
        normalize_text,
        read_monthly_accounts,
        safe_number,
        validate_month,
    )
    from .fiscal_history import apple_period
except ImportError:  # pragma: no cover
    from build_insight import (
        MONTHLY_ACCOUNT_SHEET,
        MONTHLY_NOTE_SHEET,
        _discover_monthly_workbook,
        _find_header_row,
        _identifier,
        normalize_text,
        read_monthly_accounts,
        safe_number,
        validate_month,
    )
    from fiscal_history import apple_period


METRICS = ("reads", "likes", "collects", "comments", "shares", "new_fans")


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], pattern).date()
        except ValueError:
            continue
    return None


def fiscal_label(month: str) -> str:
    period = apple_period(month)
    return f"{period['fiscal_year']} {period['fiscal_quarter']}"


def month_range(end_month: str, count: int = 12) -> list[str]:
    year, month = map(int, end_month.split("-"))
    values = []
    for _ in range(count):
        values.append(f"{year:04d}-{month:02d}")
        month -= 1
        if month == 0:
            year, month = year - 1, 12
    return list(reversed(values))


def read_note_details(path: Path) -> list[dict[str, Any]]:
    required = {
        "title": ("笔记名称",), "url": ("笔记链接",), "note_id": ("笔记ID",), "author_id": ("作者ID",),
        "category": ("笔记类型",), "format": ("笔记形式",), "published": ("笔记发布时间",),
        "reads": ("阅读次数",), "likes": ("点赞次数",), "collects": ("收藏次数",),
        "comments": ("评论次数",), "shares": ("分享次数",), "new_fans": ("单条笔记涨粉数",),
    }
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[MONTHLY_NOTE_SHEET]
        header_row, headers = _find_header_row(sheet, required)
        indexes = {field: headers.index(alias) for field, aliases in required.items() for alias in aliases if alias in headers}
        notes = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            note_id = _identifier(row[indexes["note_id"]] if len(row) > indexes["note_id"] else None)
            published = parse_date(row[indexes["published"]] if len(row) > indexes["published"] else None)
            if not note_id or not published:
                continue
            item = {
                "note_id": note_id,
                "title": _identifier(row[indexes["title"]] if len(row) > indexes["title"] else None),
                "url": _identifier(row[indexes["url"]] if len(row) > indexes["url"] else None),
                "author_id": _identifier(row[indexes["author_id"]] if len(row) > indexes["author_id"] else None),
                "category": _identifier(row[indexes["category"]] if len(row) > indexes["category"] else None) or "未分类",
                "format": _identifier(row[indexes["format"]] if len(row) > indexes["format"] else None) or "未标注",
                "published": published,
            }
            item.update({metric: round(safe_number(row[indexes[metric]] if len(row) > indexes[metric] else None)) for metric in METRICS})
            notes.append(item)
        return notes
    finally:
        workbook.close()


def compact_number(value: float | int) -> str:
    value = float(value)
    if abs(value) >= 100000000:
        return f"{value / 100000000:.2f}亿"
    if abs(value) >= 10000:
        return f"{value / 10000:.1f}万"
    return f"{value:,.0f}"


def totals(rows: list[dict[str, Any]]) -> dict[str, int]:
    result = {metric: 0 for metric in METRICS}
    result["notes"] = len(rows)
    for row in rows:
        for metric in METRICS:
            result[metric] += int(row.get(metric, 0))
    result["interactions"] = result["likes"] + result["collects"] + result["comments"]
    return result


def build_payload(month: str, data_root: Path) -> dict[str, Any]:
    month = validate_month(month)
    workbook = _discover_monthly_workbook(data_root, month)
    accounts = read_monthly_accounts(workbook, MONTHLY_ACCOUNT_SHEET, month)
    cutoff = date(int(month[:4]), int(month[5:]), __import__("calendar").monthrange(int(month[:4]), int(month[5:]))[1])
    notes = [note for note in read_note_details(workbook) if note["published"] <= cutoff]
    current_notes = [note for note in notes if note["published"].strftime("%Y-%m") == month]
    # Content diagnostics always use the selected reporting month; the full library remains for history only.
    analysis_notes = current_notes
    current_account_totals = totals(accounts)
    current_account_totals["notes"] = sum(1 for account in accounts if any(account.get(key, 0) for key in ("reads", "likes", "collects", "comments", "new_fans")))
    history_months = month_range(month)
    monthly = []
    for item_month in history_months:
        rows = [note for note in notes if note["published"].strftime("%Y-%m") == item_month]
        monthly.append({"month": item_month, **totals(rows)})
    category_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for note in analysis_notes:
        category_rows[note["category"]].append(note)
    categories = []
    for category, rows in category_rows.items():
        item = totals(rows)
        item.update({"category": category, "avg_reads": round(item["reads"] / item["notes"]) if item["notes"] else 0})
        categories.append(item)
    categories.sort(key=lambda item: item["reads"], reverse=True)
    formats = Counter(note["format"] for note in analysis_notes)
    top_notes = sorted(analysis_notes, key=lambda note: note["reads"], reverse=True)[:20]
    viral_threshold = 10000
    account_map = {account["author_id"]: account for account in accounts}
    top_accounts = sorted(accounts, key=lambda account: account["reads"], reverse=True)[:8]
    performance_rows: dict[str, dict[str, Any]] = {}
    for note in analysis_notes:
        author_id = note["author_id"]
        item = performance_rows.setdefault(author_id, {"author_id": author_id, "notes": 0, "reads": 0, "viral": 0})
        item["notes"] += 1
        item["reads"] += note["reads"]
        item["viral"] += int(note["reads"] >= viral_threshold)
    account_performance = []
    for author_id, item in performance_rows.items():
        account = account_map.get(author_id, {})
        average_reads = item["reads"] / item["notes"] if item["notes"] else 0
        cohort = "S" if average_reads >= 5000 else "A" if average_reads >= 2000 else "B" if average_reads >= 500 else "C"
        account_performance.append({
            "name": account.get("account_name") or author_id,
            "dealer": account.get("dealer", ""),
            "notes": item["notes"],
            "reads": item["reads"],
            "avg_reads": round(average_reads),
            "viral_rate": item["viral"] / item["notes"] if item["notes"] else 0,
            "cohort": cohort,
        })
    account_performance.sort(key=lambda item: item["avg_reads"], reverse=True)
    lifetime = totals(notes)
    analysis = totals(analysis_notes)
    viral_notes = sum(1 for note in analysis_notes if note["reads"] >= viral_threshold)
    latest_note_month = max((note["published"].strftime("%Y-%m") for note in notes), default=month)
    top_category = categories[0] if categories else {"category": "未分类", "reads": 0, "notes": 0, "avg_reads": 0}
    weakest_category = min((row for row in categories if row["notes"] >= 10), key=lambda row: row["avg_reads"], default=top_category)
    active_accounts = len({note["author_id"] for note in analysis_notes})
    insights = [
        f"{month} 有 {active_accounts} 个账号发布笔记，累计阅读 {compact_number(analysis['reads'])}，互动 {compact_number(analysis['interactions'])}。",
        f"{month} 发布的笔记共 {compact_number(analysis['notes'])} 条；{top_category['category']} 是当月阅读规模最大的分类，占当月阅读 {top_category['reads'] / analysis['reads']:.1%}" if analysis["reads"] else "本月笔记库暂无可用阅读数据。",
        f"{weakest_category['category']} 已发布 {compact_number(weakest_category['notes'])} 条，但单篇平均阅读为 {compact_number(weakest_category['avg_reads'])}；应优先复盘选题与封面，而不是只增加频次。",
        f"{month} 有 {compact_number(viral_notes)} 条笔记达到 {compact_number(viral_threshold)} 阅读阈值，占当月笔记 {viral_notes / analysis['notes']:.1%}" if analysis["notes"] else "暂无爆款率数据。",
    ]
    return {
        "schema_version": "1.0",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source": {"month": month, "workbook": workbook.name, "cutoff": cutoff.isoformat(), "fiscal_period": fiscal_label(month)},
        "summary": {"accounts": len(accounts), "active_accounts": active_accounts, "latest_note_month": latest_note_month, "lifetime": lifetime, "analysis_month": analysis, "current_month": current_account_totals, "current_month_notes": totals(current_notes), "viral_threshold": viral_threshold, "viral_notes": viral_notes},
        "monthly": monthly,
        "categories": categories[:8],
        "formats": [{"name": name, "count": count} for name, count in formats.most_common(6)],
        "top_notes": [{
            **{**note, "published": note["published"].isoformat()},
            "account": account_map.get(note["author_id"], {}).get("account_name", "未匹配账号"),
        } for note in top_notes],
        "top_accounts": [{"name": account["account_name"] or account["author_id"], "dealer": account["dealer"], "reads": round(account["reads"]), "interactions": round(account["likes"] + account["collects"] + account["comments"]), "new_fans": round(account["new_fans"])} for account in top_accounts],
        "account_performance": account_performance,
        "insights": insights,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a source-backed XHS matrix report dataset")
    parser.add_argument("--month", required=True, help="Source month in YYYY-MM format")
    parser.add_argument("--data-root", default="数据")
    parser.add_argument("--output", default="笔记报告/insight/generated/xhs_report_data.json")
    args = parser.parse_args()
    payload = build_payload(args.month, Path(args.data_root).resolve())
    output = Path(args.output).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Built {output} from {payload['source']['workbook']}")


if __name__ == "__main__":
    main()
