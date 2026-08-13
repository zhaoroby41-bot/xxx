from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0


def text(value: Any) -> str:
    return str(value or "").strip()


def product_from_title(title: str) -> str:
    lower = title.lower()
    for word, product in (("iphone", "iPhone"), ("mac", "Mac"), ("ipad", "iPad"), ("airpods", "AirPods"), ("watch", "Apple Watch")):
        if word in lower:
            return product
    return "iPhone"


def copy_text(product: str, copy_type: str, formula: str, hotspot: str) -> tuple[str, str, list[str], str]:
    hotspot_copy = {
        "国补截止": "趁政策窗口还在，把需求和到店体验一次确认清楚。",
        "开学季": "开学前把学习、创作和通勤场景一次想明白，后面会省很多事。",
        "iPhone17": "新品信息越来越多，先把自己真正要解决的使用问题列出来。",
        "Watch11": "运动、健康和日常提醒这类高频体验，最适合现场上手比较。",
        "双十一": "大促前先完成体验和清单，真正做决定时会更从容。",
        "无": "先把自己的高频使用场景想清楚，再决定是否需要换新。",
    }[hotspot]
    if copy_type == "product":
        titles = {
            "1": f"{product} 选错很可惜，先看这 3 点",
            "2": f"用了一周 {product}，这点最意外",
            "3": f"换 {product} 前，这一步别跳过",
            "4": f"开学前，{product} 这样选更稳",
            "1+3": f"{product} 换新前，3 个问题问自己",
        }
        body = f"很多人选 {product} 时，容易把注意力都放在参数上，却忽略了自己每天真正会用到的场景。先从通勤、学习或创作里挑出一个最常见的任务，再去比较屏幕、续航和协同体验。{hotspot_copy}\n\n建议到 Apple 授权店把关键功能逐项上手，确认后再决定。收藏这篇，到店时照着体验就好。"
        image = f"首图用 {product} 的真实使用近景，第二张用 3 点体验清单做信息图。"
    elif copy_type == "store":
        titles = {
            "1": "到 Apple 授权店，先问这 3 件事",
            "2": "第一次到店体验，别只看参数",
            "3": "换新前到店试一次，真的很重要",
            "4": "开学装备用什么？先来试一遍",
            "1+3": "到店前先准备这份体验清单",
        }
        body = f"想换 {product}，不需要急着下结论。带着你现在的设备和最常用的场景到店，把迁移、配件和日常体验一次问清楚，会比反复看参数更有效率。{hotspot_copy}\n\n评论区说说你正在用什么设备，我们把适合现场体验的功能整理给你。"
        image = "封面用门店真实上手场景，人物与设备同框；图文展示体验清单和咨询点。"
    else:
        titles = {
            "1": f"准备换 {product}？先把这件事做了",
            "2": f"{product} 值不值得换，看完再决定",
            "3": f"换新窗口到了，{product} 先这样体验",
            "4": "开学焕新，先给自己一份清单",
            "1+3": f"{product} 换新前的 3 步确认法",
        }
        body = f"换新不只是选一台设备，更是让新设备接上你的学习、工作和生活节奏。{hotspot_copy}\n\n到 Apple 授权店现场体验后，再确认最适合自己的选择。把这篇转给正在纠结的朋友，一起把体验安排上。"
        image = f"封面为 {product} 与门店场景组合，加入清晰的活动时间或场景文字，画面干净克制。"
    return titles[formula], body, [f"#{product}", "#Apple授权店", "#到店体验", "#换新攻略", "#数码好物", "#Apple体验"], image


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    workbook = load_workbook(args.input, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    column = {name: headers.index(name) for name in headers}
    notes: list[dict[str, Any]] = []
    product_counts: Counter[str] = Counter()
    for row in rows:
        title = text(row[column["笔记名称"]])
        if not title:
            continue
        product = product_from_title(title)
        product_counts[product] += 1
        notes.append({
            "title": title, "url": text(row[column["笔记链接"]]), "account": text(row[column["作者昵称"]]),
            "published": text(row[column["笔记发布时间"]]), "reads": number(row[column["累计阅读数"]]),
            "likes": number(row[column["累计点赞数"]]), "collects": number(row[column["累计收藏数"]]),
            "comments": number(row[column["累计评论数"]]), "shares": number(row[column["累计分享数"]]), "product": product,
        })
    workbook.close()
    for item in notes:
        item["engagement"] = item["likes"] + item["collects"] + item["comments"] + item["shares"]
    top = sorted(notes, key=lambda item: (item["reads"], item["engagement"]), reverse=True)[:5]
    for rank, item in enumerate(top, 1): item["rank"] = rank
    product = product_counts.most_common(1)[0][0] if product_counts else "iPhone"
    total_collects = sum(item["collects"] for item in notes)
    total_comments = sum(item["comments"] for item in notes)
    insights = [
        {"id": "pain", "name": "高收藏·真实痛点", "evidence": f"累计收藏 {total_collects:,.0f}，高表现标题多以具体问题和细节切入。", "copy": "用真实使用痛点开头，正文给出可验证的体验路径。"},
        {"id": "subsidy", "name": "国补曝光效率", "evidence": "换新、选购和到店咨询类表达更适合承接短期政策窗口。", "copy": "避免只谈价格，强调政策窗口内的到店体验与选择清单。"},
        {"id": "school", "name": "开学季收藏峰值", "evidence": "学习、创作和通勤等高频场景适合沉淀为可收藏的选购内容。", "copy": "围绕开学装备清单和场景对比，强化收藏与转发。"},
        {"id": "store", "name": "门店评论转化", "evidence": f"累计评论 {total_comments:,.0f}，咨询型内容需要明确的评论区承接。", "copy": "用“你在用什么设备”这类问题收口，促进评论后到店。"},
    ]
    hotspots = [
        {"id": "国补截止", "time": "近期", "urgency": "urgent", "type": "政策窗口", "title": "国补截止窗口", "strategy": "以换新清单、到店体验与政策确认做内容承接。"},
        {"id": "开学季", "time": "8月-9月初", "urgency": "near", "type": "季节场景", "title": "开学季", "strategy": "围绕学习、创作、宿舍与通勤场景做可收藏内容。"},
        {"id": "iPhone18", "time": "9月上旬", "urgency": "near", "type": "新品预热", "title": "iPhone 18 预热", "strategy": "以需求梳理和到店体验为主，避免未证实参数。"},
        {"id": "WatchS12", "time": "9月上旬", "urgency": "near", "type": "新品预热", "title": "Apple Watch S12", "strategy": "突出运动、健康与日常提醒的真实体验。"},
        {"id": "AirPods", "time": "9月", "urgency": "later", "type": "新品预热", "title": "AirPods 新款", "strategy": "围绕通勤、学习与生态协同做场景化预热。"},
        {"id": "双十一", "time": "10月-11月", "urgency": "later", "type": "大促准备", "title": "双十一", "strategy": "提前沉淀选购、对比和到店体验内容，放大收藏价值。"},
    ]
    formula_defs = [
        {"id": "1", "name": "痛点数字型", "sample": top[4] if len(top) > 4 else top[0], "hook": "具体问题 + 数字承诺", "rhythm": "痛点 → 3 点判断 → 体验验证 → 收藏 CTA", "cta": "收藏体验清单", "emoji": "低密度专业感", "keywords": "屏幕、误区、怎么选"},
        {"id": "2", "name": "体验分享型", "sample": top[1] if len(top) > 1 else top[0], "hook": "真实感受 + 反差", "rhythm": "体验感受 → 场景说明 → 建议 → 评论 CTA", "cta": "评论说设备", "emoji": "生活化点缀", "keywords": "真实体验、日常、值不值"},
        {"id": "3", "name": "活动紧迫型", "sample": top[0], "hook": "热点窗口 + 行动提醒", "rhythm": "时间节点 → 需求梳理 → 到店体验 → 行动 CTA", "cta": "预约到店", "emoji": "适度强调", "keywords": "换新、窗口、到店"},
        {"id": "4", "name": "开学种草型", "sample": top[2] if len(top) > 2 else top[0], "hook": "开学场景 + 想象结果", "rhythm": "场景代入 → 清单 → 使用结果 → 收藏 CTA", "cta": "收藏开学清单", "emoji": "轻量生活化", "keywords": "开学、学习、通勤"},
    ]
    formulas = []
    for definition in formula_defs:
        sample = definition["sample"]
        formulas.append({**{key: value for key, value in definition.items() if key != "sample"}, "sample": sample, "title_template": f"[{definition['name']}] + [具体场景] + [行动提示]", "effect": f"来源样本累计阅读 {sample['reads']:,.0f}"})
    type_specs = [("product", "产品种草", 7), ("store", "门店体验", 5), ("promo", "活动促销", 6)]
    formula_ids = ["1", "2", "3", "4", "1+3"]
    insight_ids = ["pain", "subsidy", "school", "store"]
    hotspot_ids = ["国补截止", "开学季", "iPhone17", "Watch11", "双十一", "无"]
    themes = [
        ("festival", "节日营销"),
        ("campaign", "营销活动"),
        ("promotion", "促销优惠"),
        ("system", "系统功能"),
        ("tips", "使用技巧"),
        ("product", "产品种草"),
        ("other", "其他"),
    ]
    copies = []
    cursor = 0
    for type_id, type_name, count in type_specs:
        for _ in range(count):
            formula = formula_ids[cursor % len(formula_ids)]
            insight = insights[cursor % len(insights)]
            hotspot = hotspot_ids[cursor % len(hotspot_ids)]
            theme_id, theme_name = themes[cursor % len(themes)]
            if hotspot in {"国补截止", "双十一"}:
                theme_id, theme_name = "promotion", "促销优惠"
            elif hotspot == "开学季":
                theme_id, theme_name = "festival", "节日营销"
            elif type_id == "store":
                theme_id, theme_name = "campaign", "营销活动"
            title, body, tags, image = copy_text(product, type_id, formula, hotspot)
            copies.append({"id": cursor + 1, "type": type_id, "type_name": type_name, "theme": theme_id, "theme_name": theme_name, "formula": formula, "formula_name": next((item["name"] for item in formulas if item["id"] == formula), "痛点数字型 + 活动紧迫型"), "insight": insight["id"], "insight_name": insight["name"], "hotspot": hotspot, "title": f"{['📱', '🏪', '🎁'][['product', 'store', 'promo'].index(type_id)]} {title}", "body": body, "tags": tags, "image_direction": image})
            cursor += 1
    payload = {"generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"), "source": args.input.name, "summary": {"notes": len(notes), "top_samples": len(top), "formulas": len(formulas), "copies": len(copies), "types": len(type_specs)}, "samples": top, "insights": insights, "hotspots": hotspots, "formulas": formulas, "copies": copies}
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Generated {len(copies)} copies from {len(notes)} notes -> {args.output}")


if __name__ == "__main__":
    main()
