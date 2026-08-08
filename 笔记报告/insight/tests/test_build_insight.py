import sys
import unittest
import json
import tempfile
import subprocess
from pathlib import Path

from openpyxl import Workbook


SOURCE_ROOT = Path(__file__).resolve().parents[1]
if str(SOURCE_ROOT) not in sys.path:
    sys.path.insert(0, str(SOURCE_ROOT))

import scripts.build_insight as build_insight

from scripts.build_insight import (
    calculate_content_metrics,
    calculate_kpi_status,
    classify_account_cohort,
    map_category,
    normalize_text,
    resolve_region,
    safe_number,
)


class MetricTests(unittest.TestCase):
    def test_safe_number_defaults_invalid_values_to_zero(self):
        self.assertEqual(safe_number(None), 0.0)
        self.assertEqual(safe_number("not-a-number"), 0.0)
        self.assertEqual(safe_number("12.5"), 12.5)

    def test_normalize_text_collapses_whitespace(self):
        self.assertEqual(normalize_text("  ONEZERO\n"), "ONEZERO")

    def test_kpi_status_uses_apple_q4_pacing(self):
        result = calculate_kpi_status(400, 1000, 1 / 3)
        self.assertAlmostEqual(result["completion_rate"], 0.4)
        self.assertAlmostEqual(result["pacing_gap"], 0.4 - 1 / 3)
        self.assertEqual(result["status"], "normal")

    def test_kpi_status_marks_leading_at_exact_ten_point_gap(self):
        result = calculate_kpi_status(40, 100, 0.3)
        self.assertEqual(result["status"], "leading")

    def test_kpi_status_marks_warning_at_exact_five_point_gap(self):
        result = calculate_kpi_status(25, 100, 0.3)
        self.assertEqual(result["status"], "normal")
        result = calculate_kpi_status(24, 100, 0.3)
        self.assertEqual(result["status"], "warning")

    def test_kpi_status_marks_critical_below_exact_fifteen_point_gap(self):
        result = calculate_kpi_status(15, 100, 0.3)
        self.assertEqual(result["status"], "warning")
        result = calculate_kpi_status(14, 100, 0.3)
        self.assertEqual(result["status"], "critical")

    def test_kpi_status_marks_leading_warning_and_critical(self):
        self.assertEqual(calculate_kpi_status(500, 1000, 1 / 3)["status"], "leading")
        self.assertEqual(calculate_kpi_status(200, 1000, 1 / 3)["status"], "warning")
        self.assertEqual(calculate_kpi_status(100, 1000, 1 / 3)["status"], "critical")

    def test_kpi_status_rejects_missing_target(self):
        result = calculate_kpi_status(400, 0, 1 / 3)
        self.assertIsNone(result["completion_rate"])
        self.assertEqual(result["status"], "unmatched")

    def test_content_metrics_use_defined_interactions(self):
        result = calculate_content_metrics({
            "reads": 1000,
            "likes": 50,
            "collects": 20,
            "comments": 10,
            "shares": 99,
            "new_fans": 8,
            "visitors": 40,
            "notes": 5,
        })
        self.assertEqual(result["interactions"], 80)
        self.assertAlmostEqual(result["interaction_rate"], 0.08)
        self.assertAlmostEqual(result["fans_per_10k_reads"], 80)
        self.assertAlmostEqual(result["reads_per_note"], 200)


class InsightModelInterfaceTests(unittest.TestCase):
    def test_task_four_public_builders_are_available(self):
        required = (
            "build_dealer_insights",
            "build_apple_insights",
            "build_insight_payload",
        )
        self.assertTrue(all(hasattr(build_insight, name) for name in required))


class InsightModelTests(unittest.TestCase):
    CATEGORY_MAPPING = {
        "A": {"unified": "A", "confirmed": True},
        "B": {"unified": "B", "confirmed": True},
    }

    def sample_inputs(self):
        accounts = [
            {"dealer": "Dealer A", "store": "Beijing Store", "account_name": "Core A", "author_id": "a1", "xhs_id": "x1", "reads": 100, "likes": 10, "collects": 10, "comments": 10, "shares": 100, "new_fans": 4, "visitors": 5},
            {"dealer": "Dealer A", "store": "Shanghai Store", "account_name": "Expanded A", "author_id": "a2", "xhs_id": "x2", "reads": 500, "likes": 20, "collects": 20, "comments": 20, "shares": 200, "new_fans": 10, "visitors": 30},
            {"dealer": "Dealer B", "store": "Beijing Store", "account_name": "Core B", "author_id": "b1", "xhs_id": "x3", "reads": 200, "likes": 20, "collects": 20, "comments": 20, "shares": 300, "new_fans": 8, "visitors": 10},
        ]
        notes = [
            {"note_id": "july-1", "author_id": "a1", "publish_date": "2026-07-03", "note_format": "image", "category": "A", "reads": 100, "likes": 10, "collects": 10, "comments": 10, "shares": 99, "new_fans": 4},
            {"note_id": "old-1", "author_id": "a1", "publish_date": "2026-06-30", "note_format": "video", "category": "A", "reads": 9999, "likes": 999, "collects": 999, "comments": 999, "shares": 999, "new_fans": 999},
            {"note_id": "july-2", "author_id": "b1", "publish_date": "2026-07-04", "note_format": "video", "category": "B", "reads": 200, "likes": 20, "collects": 20, "comments": 20, "shares": 88, "new_fans": 8},
        ]
        kpis = [
            {"author_id": "a1", "account_name": "Core A", "read_target": 300, "interaction_target": 90, "fan_target": 12},
            {"author_id": "b1", "account_name": "Core B", "read_target": 300, "interaction_target": 90, "fan_target": 12},
        ]
        return accounts, notes, kpis

    def test_dealer_grouping_aggregates_children_and_isolates_core_kpi_metrics(self):
        accounts, notes, kpis = self.sample_inputs()
        dealers = build_insight.build_dealer_insights(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, "2026-07")

        dealer = next(item for item in dealers if item["name"] == "Dealer A")
        self.assertEqual(dealer["cohort"], "core_kpi")
        self.assertEqual(len(dealer["accounts"]), 2)
        self.assertEqual(dealer["kpi"]["reads"]["actual"], 100)
        self.assertEqual(dealer["kpi"]["reads"]["target"], 300)
        self.assertEqual(dealer["expanded_store_metrics"]["reads"], 500)

    def test_interactions_exclude_shares_and_month_filter_excludes_old_notes(self):
        accounts, notes, kpis = self.sample_inputs()
        dealer = build_insight.build_dealer_insights(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, "2026-07")[0]

        self.assertEqual(dealer["content"]["notes"], 1)
        self.assertEqual(dealer["content"]["reads"], 100)
        self.assertEqual(dealer["content"]["interactions"], 30)
        self.assertEqual(dealer["content"]["shares"], 99)
        self.assertEqual(dealer["kpi"]["interactions"]["actual"], 30)

    def test_q4_elapsed_ratio_uses_completed_months(self):
        self.assertEqual(build_insight.q4_elapsed_ratio("2026-07"), 1 / 3)
        self.assertEqual(build_insight.q4_elapsed_ratio("2026-08"), 2 / 3)
        self.assertEqual(build_insight.q4_elapsed_ratio("2026-09"), 1)

    def test_category_rules_cover_supported_signal_and_validate_states(self):
        categories = [
            {"category": "A", "notes": 3, "note_share": 0.1, "reads_per_note": 200, "benchmark_note_share": 0.3, "benchmark_reads_per_note": 100, "benchmark_confidence": "supported"},
            {"category": "B", "notes": 3, "note_share": 0.5, "reads_per_note": 50, "benchmark_note_share": 0.3, "benchmark_reads_per_note": 100, "benchmark_confidence": "supported"},
            {"category": "C", "notes": 3, "note_share": 0.5, "reads_per_note": 200, "benchmark_note_share": 0.3, "benchmark_reads_per_note": 100, "benchmark_confidence": "supported"},
            {"category": "D", "notes": 1, "note_share": 0.1, "reads_per_note": 200, "benchmark_note_share": 0.3, "benchmark_reads_per_note": 100, "benchmark_confidence": "validate"},
        ]
        recommendations = build_insight.generate_category_recommendations("dealer-a", categories)
        self.assertEqual({item["confidence"] for item in recommendations}, {"supported", "signal", "validate"})
        self.assertEqual({item["target"]["category"] for item in recommendations}, {"A", "B", "C", "D"})
        self.assertEqual(
            {item["rule_id"] for item in recommendations},
            {"category_scale", "category_optimize", "category_maintain", "category_validate"},
        )

    def test_recommendations_have_evidence_and_allowed_enums(self):
        accounts, notes, kpis = self.sample_inputs()
        dealers = build_insight.build_dealer_insights(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, "2026-07")
        recommendations = [item for dealer in dealers for item in dealer["recommendations"]]
        self.assertTrue(recommendations)
        for recommendation in recommendations:
            self.assertTrue(recommendation["evidence"])
            self.assertIn(recommendation["confidence"], {"supported", "signal", "validate"})
            self.assertIn(recommendation["priority"], {"high", "medium", "low"})

    def test_apple_completion_is_target_weighted_and_statuses_reconcile(self):
        dealers = [
            {"dealer_id": "a", "name": "A", "cohort": "core_kpi", "accounts": [{"cohort": "core_kpi"}], "kpi": {"overall_status": "leading", "reads": {"actual": 90, "target": 100, "completion_rate": .9, "pacing_gap": .5}, "interactions": {"actual": 9, "target": 10, "completion_rate": .9, "pacing_gap": .5}, "fans": {"actual": 9, "target": 10, "completion_rate": .9, "pacing_gap": .5}}, "expanded_store_metrics": {"reads": 0}, "content": {"categories": []}, "recommendations": []},
            {"dealer_id": "b", "name": "B", "cohort": "core_kpi", "accounts": [{"cohort": "core_kpi"}, {"cohort": "expanded_store"}], "kpi": {"overall_status": "warning", "reads": {"actual": 1, "target": 900, "completion_rate": 1 / 900, "pacing_gap": -.32}, "interactions": {"actual": 1, "target": 90, "completion_rate": 1 / 90, "pacing_gap": -.23}, "fans": {"actual": 1, "target": 90, "completion_rate": 1 / 90, "pacing_gap": -.23}}, "expanded_store_metrics": {"reads": 500}, "content": {"categories": []}, "recommendations": []},
        ]
        apple = build_insight.build_apple_insights(dealers, {"matched_kpi_accounts": 2, "unmatched_kpi_ids": ["missing"]}, "2026-07")

        self.assertAlmostEqual(apple["network_kpis"]["reads"]["completion_rate"], 91 / 1000)
        self.assertNotAlmostEqual(apple["network_kpis"]["reads"]["completion_rate"], (.9 + 1 / 900) / 2)
        self.assertEqual(sum(apple["status_counts"][key] for key in ("leading", "normal", "warning", "critical")), 2)
        self.assertEqual(apple["account_counts"], {"core_kpi": 2, "expanded_store": 1})

    def test_risk_dealers_prioritize_severity_then_worst_pacing_gap(self):
        def dealer(dealer_id, name, status, gap):
            metric = {"actual": 1, "target": 100, "completion_rate": .01, "pacing_gap": gap, "status": status}
            return {
                "dealer_id": dealer_id, "name": name, "cohort": "core_kpi",
                "accounts": [{"author_id": f"{dealer_id}-account", "cohort": "core_kpi", "region": "\u534e\u4e1c", "city": "\u4e0a\u6d77", "confidence": "confirmed"}],
                "kpi": {"overall_status": status, "account_statuses": [{"status": status}], "reads": metric, "interactions": metric, "fans": metric},
                "content": {"notes": 1, "reads": 1, "reads_per_note": 1, "categories": []},
                "recommendations": [],
            }

        apple = build_insight.build_apple_insights([
            dealer("warning", "Warning Dealer", "warning", -.10),
            dealer("critical-less", "Critical Less", "critical", -.20),
            dealer("critical-worse", "Critical Worse", "critical", -.50),
        ], {"matched_kpi_accounts": 3, "unmatched_kpi_ids": []}, "2026-07")

        self.assertEqual(
            [item["dealer_id"] for item in apple["risk_dealers"]],
            ["critical-worse", "critical-less", "warning"],
        )
        self.assertEqual(
            [item.get("worst_pacing_gap") for item in apple["risk_dealers"]],
            [-.50, -.20, -.10],
        )

    def test_payload_is_deterministic_except_time_and_rejects_non_finite_values(self):
        accounts, notes, kpis = self.sample_inputs()
        quality = {"source_files": {}, "matched_kpi_accounts": 2, "unmatched_kpi_ids": [], "generated_at": "2026-08-06T00:00:00Z"}
        first = build_insight.build_insight_payload(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, quality, "2026-07")
        second = build_insight.build_insight_payload(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, quality, "2026-07")
        first.pop("generated_at")
        second.pop("generated_at")
        self.assertEqual(first, second)
        self.assertNotIn("generated_at", first["quality"])
        json.dumps(first, allow_nan=False)

    def test_dealer_scoped_builders_emit_index_and_one_dealer_payloads_only(self):
        accounts, notes, kpis = self.sample_inputs()
        quality = {
            "source_files": {},
            "matched_kpi_accounts": 2,
            "unmatched_kpi_ids": [],
            "quality_status": "ready",
            "errors": [],
            "data_freshness": {
                "source_snapshot_at": "2026-08-01",
                "basis": "note_export_timestamp",
                "is_fallback": False,
            },
        }
        payload = build_insight.build_insight_payload(
            accounts, notes, kpis, self.CATEGORY_MAPPING, {}, quality, "2026-07"
        )

        index = build_insight.build_dealer_index(payload)
        self.assertEqual(
            set(index),
            {"schema_version", "source_month", "generated_at", "data_freshness", "quality", "dealers"},
        )
        self.assertEqual(
            index["dealers"],
            [{"dealer_id": item["dealer_id"], "name": item["name"]} for item in payload["dealers"]],
        )
        self.assertEqual(index["quality"], {"quality_status": "ready", "errors": []})
        self.assertNotIn("accounts", json.dumps(index, ensure_ascii=False))
        self.assertNotIn("apple", index)

        selected = payload["dealers"][0]
        other = payload["dealers"][1]
        scoped = build_insight.build_dealer_payload(payload, selected["dealer_id"])
        self.assertEqual(
            set(scoped),
            {"schema_version", "source_month", "period", "generated_at", "data_freshness", "quality", "history", "dealer"},
        )
        self.assertEqual(scoped["dealer"], selected)
        serialized = json.dumps(scoped, ensure_ascii=False)
        self.assertNotIn(other["dealer_id"], serialized)
        self.assertNotIn(other["name"], serialized)
        self.assertNotIn("apple", scoped)
        self.assertNotIn("dealers", scoped)
        self.assertNotIn("unmatched_kpi_ids", serialized)

    def test_write_dealer_scoped_artifacts_creates_expected_routes(self):
        accounts, notes, kpis = self.sample_inputs()
        payload = build_insight.build_insight_payload(
            accounts,
            notes,
            kpis,
            self.CATEGORY_MAPPING,
            {},
            {
                "source_files": {},
                "matched_kpi_accounts": 2,
                "unmatched_kpi_ids": [],
                "quality_status": "ready_with_warnings",
                "errors": [],
                "data_freshness": {
                    "source_snapshot_at": "2026-08-01",
                    "basis": "note_export_timestamp",
                    "is_fallback": False,
                },
            },
            "2026-07",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            stale_path = output_dir / "dealers" / "dealer-stale.json"
            stale_path.parent.mkdir(parents=True)
            stale_path.write_text('{"dealer":{"dealer_id":"dealer-stale"}}', encoding="utf-8")
            build_insight.write_dealer_scoped_artifacts(output_dir, payload)

            index = json.loads((output_dir / "dealer_index.json").read_text(encoding="utf-8"))
            self.assertEqual(len(index["dealers"]), 2)
            self.assertFalse(stale_path.exists())
            self.assertEqual(
                sorted(path.stem for path in (output_dir / "dealers").glob("*.json")),
                sorted(item["dealer_id"] for item in payload["dealers"]),
            )
            for dealer in payload["dealers"]:
                scoped = json.loads(
                    (output_dir / "dealers" / f"{dealer['dealer_id']}.json").read_text(encoding="utf-8")
                )
                self.assertEqual(scoped["dealer"]["dealer_id"], dealer["dealer_id"])

    def test_versioned_dealer_artifact_contains_no_peer_identity(self):
        accounts, notes, kpis = self.sample_inputs()
        payload = build_insight.build_insight_payload(
            accounts,
            notes,
            kpis,
            self.CATEGORY_MAPPING,
            {},
            {
                "source_files": {},
                "matched_kpi_accounts": 2,
                "unmatched_kpi_ids": [],
                "quality_status": "ready",
                "errors": [],
            },
            "2026-07",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            build_insight.attach_ai_artifacts(payload, output_dir, use_provider=False)
            build_insight.write_versioned_artifacts(output_dir, payload)

            selected = payload["dealers"][0]
            peer = payload["dealers"][1]
            artifact = json.loads(
                (output_dir / "months" / "2026-07" / "dealers" / f"{selected['dealer_id']}.json").read_text(encoding="utf-8")
            )
            serialized = json.dumps(artifact, ensure_ascii=False)
            self.assertIn(selected["dealer_id"], serialized)
            self.assertNotIn(peer["dealer_id"], serialized)
            self.assertNotIn(peer["name"], serialized)
            self.assertIn("ai_insights", artifact["dealer"])
            self.assertIn("evidence", artifact["dealer"])

    def test_apple_month_artifact_has_period_history_evidence_and_ai(self):
        accounts, notes, kpis = self.sample_inputs()
        payload = build_insight.build_insight_payload(
            accounts,
            notes,
            kpis,
            self.CATEGORY_MAPPING,
            {},
            {"source_files": {}, "matched_kpi_accounts": 2, "unmatched_kpi_ids": [], "quality_status": "ready", "errors": []},
            "2026-07",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            build_insight.attach_ai_artifacts(payload, output_dir, use_provider=False)
            build_insight.write_versioned_artifacts(output_dir, payload)

            apple = json.loads((output_dir / "months" / "2026-07" / "apple.json").read_text(encoding="utf-8"))
            self.assertEqual(apple["schema_version"], "2.0")
            self.assertEqual(apple["period"]["fiscal_quarter"], "Q4")
            self.assertIn("history", apple)
            self.assertIn("evidence", apple["apple"])
            self.assertIn("ai_insights", apple["apple"])
            self.assertEqual(apple["metadata"]["logic_version"], "2026-08-08.1")

    def test_apple_category_performance_is_segmented_by_region_and_cohort(self):
        accounts, notes, kpis = self.sample_inputs()
        payload = build_insight.build_insight_payload(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, {"source_files": {}, "matched_kpi_accounts": 2, "unmatched_kpi_ids": []}, "2026-07")
        categories = payload["apple"]["category_mix_performance"]
        self.assertTrue(categories)
        self.assertTrue(all("region" in item and "cohort" in item for item in categories))

    def test_mixed_dealer_category_rollups_use_joined_account_cohort(self):
        accounts = [
            {"dealer": "\u4e0a\u6d77\u7ecf\u9500\u5546", "store": "\u5317\u4eac\u95e8\u5e97", "account_name": "\u5317\u4eac\u6838\u5fc3\u8d26\u53f7", "author_id": "core", "reads": 100, "likes": 10, "collects": 10, "comments": 10, "new_fans": 1},
            {"dealer": "\u4e0a\u6d77\u7ecf\u9500\u5546", "store": "\u82cf\u5dde\u95e8\u5e97", "account_name": "\u82cf\u5dde\u6269\u5c55\u8d26\u53f7", "author_id": "expanded", "reads": 100, "likes": 10, "collects": 10, "comments": 10, "new_fans": 1},
        ]
        notes = [
            {"note_id": "core-note", "author_id": "core", "publish_date": "2026-07-01", "note_format": "image", "category": "A", "reads": 100, "likes": 10, "collects": 10, "comments": 10, "new_fans": 1},
            {"note_id": "expanded-note", "author_id": "expanded", "publish_date": "2026-07-01", "note_format": "image", "category": "A", "reads": 100, "likes": 10, "collects": 10, "comments": 10, "new_fans": 1},
        ]
        kpis = [{"author_id": "core", "read_target": 300, "interaction_target": 90, "fan_target": 3}]
        payload = build_insight.build_insight_payload(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, {"source_files": {}, "matched_kpi_accounts": 1, "unmatched_kpi_ids": []}, "2026-07")
        category_rows = [item for item in payload["apple"]["category_mix_performance"] if item["category"] == "A"]

        self.assertEqual({item["cohort"] for item in category_rows}, {"core_kpi", "expanded_store"})
        self.assertEqual(
            {(item["region"], item["cohort"]) for item in category_rows},
            {("\u534e\u5317", "core_kpi"), ("\u534e\u4e1c", "expanded_store")},
        )

    def test_mixed_dealer_core_quadrant_excludes_expanded_content(self):
        accounts = [
            {"dealer": "Mixed Dealer", "store": "Core Store", "account_name": "Core", "author_id": "core", "reads": 100, "likes": 1, "collects": 1, "comments": 1, "new_fans": 1},
            {"dealer": "Mixed Dealer", "store": "Expanded Store", "account_name": "Expanded", "author_id": "expanded", "reads": 9000, "likes": 90, "collects": 90, "comments": 90, "new_fans": 90},
        ]
        notes = [{
            "note_id": "core-note", "author_id": "core", "publish_date": "2026-07-01",
            "note_format": "image", "category": "A", "reads": 100, "likes": 1, "collects": 1, "comments": 1, "new_fans": 1,
        }] + [{
            "note_id": f"expanded-{index}", "author_id": "expanded", "publish_date": "2026-07-01",
            "note_format": "image", "category": "A", "reads": 1000, "likes": 10, "collects": 10, "comments": 10, "new_fans": 10,
        } for index in range(9)]
        kpis = [{"author_id": "core", "read_target": 300, "interaction_target": 90, "fan_target": 3}]

        payload = build_insight.build_insight_payload(
            accounts, notes, kpis, self.CATEGORY_MAPPING, {},
            {"source_files": {}, "matched_kpi_accounts": 1, "unmatched_kpi_ids": []}, "2026-07",
        )
        point = payload["apple"]["dealer_quadrants"][0]

        self.assertEqual(point["cohort"], "core_kpi")
        self.assertEqual(point.get("notes"), 1)
        self.assertEqual(point.get("reads"), 100)
        self.assertEqual(point.get("reads_per_note"), 100)
        self.assertEqual(point.get("source_account_ids"), ["core"])

    def test_fallback_benchmark_uses_full_same_cohort_pool(self):
        dealers = [
            {"dealer_id": "a", "cohort": "core_kpi", "content": {"reads_per_note": 1000, "categories": [{"category": "A", "note_share": .2, "reads_per_note": 100}] }},
            {"dealer_id": "b", "cohort": "core_kpi", "content": {"reads_per_note": 2000, "categories": [{"category": "A", "note_share": .4, "reads_per_note": 200}] }},
            {"dealer_id": "c", "cohort": "core_kpi", "content": {"reads_per_note": 3000, "categories": []}},
        ]
        build_insight._attach_category_benchmarks(dealers)
        category = dealers[0]["content"]["categories"][0]

        self.assertEqual(category["benchmark_confidence"], "lower_confidence")
        self.assertEqual(category["benchmark_reads_per_note"], 2000)
        self.assertNotEqual(category["benchmark_reads_per_note"], 150)

    def test_unconfirmed_mapping_never_emits_performance_strategy(self):
        recommendations = build_insight.generate_category_recommendations("dealer-a", [{
            "category": "unconfirmed", "notes": 5, "note_share": .1, "reads_per_note": 200,
            "benchmark_note_share": .3, "benchmark_reads_per_note": 100,
            "benchmark_confidence": "supported", "mapping_completeness": .8,
        }])

        self.assertEqual(len(recommendations), 1)
        self.assertEqual(recommendations[0]["type"], "data_quality")
        self.assertEqual(recommendations[0]["confidence"], "validate")
        self.assertEqual(recommendations[0]["target"]["category"], "")

    def test_apple_actions_are_bounded_aggregates_with_drilldown_references(self):
        dealers = []
        for index in range(20):
            dealers.append({
                "dealer_id": f"dealer-{index}", "name": f"Dealer {index}", "cohort": "core_kpi",
                "accounts": [{"author_id": f"account-{index}", "cohort": "core_kpi", "region": "\u534e\u4e1c", "confidence": "confirmed"}],
                "kpi": {"overall_status": "normal", "account_statuses": [], "reads": {}, "interactions": {}, "fans": {}},
                "content": {"notes": 3, "reads_per_note": 100, "categories": []},
                "recommendations": [{
                    "id": f"rec-{index}", "type": "category", "title": "\u503c\u5f97\u52a0\u7801A", "action": "\u589e\u52a0A\u4f9b\u7ed9", "confidence": "supported", "priority": "medium",
                    "evidence": [{"metric": "reads_per_note", "value": 200, "benchmark": 100, "scope": "same_cohort"}],
                    "target": {"category": "A", "city": "", "account_id": ""},
                }],
            })
        actions = build_insight.build_apple_insights(dealers, {"matched_kpi_accounts": 20, "unmatched_kpi_ids": []}, "2026-07")["actions"]
        flattened = [action for bucket in actions.values() for action in bucket]

        self.assertLessEqual(len(flattened), 36)
        self.assertEqual(len(flattened), 1)
        self.assertEqual(flattened[0]["affected_dealer_count"], 20)
        self.assertEqual(len(flattened[0]["drilldown_recommendation_ids"]), 20)
        self.assertEqual(len({action["id"] for action in flattened}), len(flattened))

    def test_replicable_cases_require_above_benchmark_and_sample(self):
        dealers = [{
            "dealer_id": "a", "name": "A", "cohort": "core_kpi", "accounts": [],
            "kpi": {"account_statuses": []},
            "content": {"notes": 3, "reads_per_note": 100, "categories": [{
                "category": "A", "notes": 3, "reads_per_note": 100, "note_share": .2,
                "benchmark_reads_per_note": 200, "benchmark_note_share": .1,
                "benchmark_confidence": "supported",
            }]},
        }]
        cases = build_insight._replicable_cases(dealers, [{"cohort": "core_kpi", "category": "A", "notes": 3, "reads_per_note": 100}])
        self.assertEqual(cases["category"], [])
        self.assertEqual(cases["city"], [])

    def test_replicable_category_requires_complete_mapping(self):
        dealers = [{
            "dealer_id": "a", "name": "A", "cohort": "core_kpi", "accounts": [], "kpi": {"account_statuses": []},
            "content": {"notes": 5, "reads_per_note": 300, "categories": [{
                "category": "A", "notes": 5, "reads_per_note": 300, "note_share": .4,
                "benchmark_reads_per_note": 100, "benchmark_note_share": .2,
                "benchmark_confidence": "supported", "mapping_completeness": .8,
            }]},
        }]
        self.assertEqual(build_insight._replicable_cases(dealers, [])["category"], [])

    def test_apple_actions_keep_semantic_rules_separate_and_use_network_copy(self):
        rules = [
            ("category_scale", "category", "Dealer scale title", "scale-1"),
            ("category_optimize", "category", "Dealer optimize title", "optimize-1"),
            ("category_maintain", "category", "Dealer maintain title", "maintain-1"),
            ("category_validate", "category", "Dealer validate title", "validate-1"),
            ("kpi_reads_critical", "kpi", "Dealer KPI title", "kpi-1"),
            ("quality_unknown_city", "data_quality", "Dealer quality title", "quality-city"),
            ("quality_unmatched_kpi", "data_quality", "Dealer unmatched title", "quality-kpi"),
        ]
        dealers = []
        for index, (rule_id, recommendation_type, title, recommendation_id) in enumerate(rules):
            dealers.append({
                "dealer_id": f"dealer-{index}", "name": f"Dealer {index}", "cohort": "core_kpi",
                "accounts": [{"author_id": f"account-{index}", "cohort": "core_kpi", "region": "\u534e\u4e1c", "confidence": "confirmed"}],
                "kpi": {"overall_status": "normal", "account_statuses": [], "reads": {}, "interactions": {}, "fans": {}},
                "content": {"notes": 3, "reads_per_note": 100, "categories": []},
                "recommendations": [{
                    "id": recommendation_id, "rule_id": rule_id, "type": recommendation_type, "title": title, "action": "Dealer action",
                    "confidence": "validate" if "validate" in rule_id or "quality" in rule_id else "supported", "priority": "high" if rule_id.startswith("kpi_") else "medium",
                    "evidence": [{"metric": "test", "value": 1, "benchmark": 0, "scope": "dealer"}],
                    "target": {"category": "A" if recommendation_type == "category" else "", "city": "", "account_id": ""},
                }],
            })
        actions = [item for bucket in build_insight.build_apple_insights(dealers, {"matched_kpi_accounts": 7, "unmatched_kpi_ids": []}, "2026-07")["actions"].values() for item in bucket]
        by_rule = {item["rule_id"]: item for item in actions}

        self.assertEqual(set(by_rule), {item[0] for item in rules})
        self.assertTrue(all(item["title"].startswith("\u7f51\u7edc") for item in actions))
        self.assertNotEqual(by_rule["kpi_reads_critical"]["title"], "Dealer KPI title")
        self.assertEqual(by_rule["kpi_reads_critical"]["affected_dealer_count"], 1)
        self.assertEqual(by_rule["category_scale"]["drilldown_recommendation_ids"], ["scale-1"])

    def test_city_replicable_uses_city_note_evidence_not_dealer_totals(self):
        accounts = [
            {"dealer": "\u4e0a\u6d77\u7ecf\u9500\u5546", "store": "\u5317\u4eac\u95e8\u5e97", "account_name": "\u5317\u4eac\u8d26\u53f7", "author_id": "beijing", "reads": 900, "likes": 90, "collects": 0, "comments": 0, "new_fans": 9},
            {"dealer": "\u4e0a\u6d77\u7ecf\u9500\u5546", "store": "\u82cf\u5dde\u95e8\u5e97", "account_name": "\u82cf\u5dde\u8d26\u53f7", "author_id": "suzhou", "reads": 300, "likes": 30, "collects": 0, "comments": 0, "new_fans": 3},
        ]
        notes = [
            {"note_id": f"beijing-{index}", "author_id": "beijing", "publish_date": "2026-07-01", "note_format": "image", "category": "A", "reads": 300, "likes": 30, "collects": 0, "comments": 0, "new_fans": 3}
            for index in range(3)
        ] + [
            {"note_id": f"suzhou-{index}", "author_id": "suzhou", "publish_date": "2026-07-01", "note_format": "image", "category": "A", "reads": 100, "likes": 10, "collects": 0, "comments": 0, "new_fans": 1}
            for index in range(3)
        ]
        kpis = [
            {"author_id": "beijing", "read_target": 900, "interaction_target": 90, "fan_target": 9},
            {"author_id": "suzhou", "read_target": 300, "interaction_target": 30, "fan_target": 3},
        ]
        payload = build_insight.build_insight_payload(accounts, notes, kpis, self.CATEGORY_MAPPING, {}, {"source_files": {}, "matched_kpi_accounts": 2, "unmatched_kpi_ids": []}, "2026-07")
        city_cases = payload["apple"]["replicable_cases"]["city"]

        self.assertEqual([item["city"] for item in city_cases], ["\u5317\u4eac"])
        self.assertEqual(city_cases[0]["evidence"][0], {"metric": "reads_per_note", "value": 300, "benchmark": 200, "scope": "city_content"})


class StandardizationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        config_root = SOURCE_ROOT / "config"
        with (config_root / "category_mapping.json").open(encoding="utf-8") as handle:
            cls.category_mapping = json.load(handle)

    def test_known_category_mapping_returns_unified_category_and_confirmation(self):
        self.assertEqual(map_category("iPhone 产品", self.category_mapping), ("产品种草", True))

    def test_empty_category_is_unclassified(self):
        self.assertEqual(map_category("  ", self.category_mapping), ("未分类", False))

    def test_unknown_category_defaults_to_other_without_confirmation(self):
        self.assertEqual(map_category("从未出现的分类", self.category_mapping), ("其他", False))

    def test_category_spelling_variants_map_identically(self):
        expected = {
            "iPhone产品": "产品种草",
            "mac 产品": "产品种草",
            "IP活动": "营销活动",
            "beats 内容": "产品种草",
        }
        for category, unified in expected.items():
            self.assertEqual(map_category(category, self.category_mapping), (unified, True))

    def test_account_override_is_authoritative(self):
        overrides = {
            "上海旗舰店": {"city": "北京", "province": "北京市", "region": "华北"}
        }
        self.assertEqual(
            resolve_region("上海经销商", "上海门店", "上海旗舰店", overrides),
            {"city": "北京", "province": "北京市", "region": "华北", "confidence": "confirmed"},
        )

    def test_author_id_override_applies_without_matching_account_name(self):
        overrides = {
            "123456": {"city": "深圳", "province": "广东省", "region": "华南"}
        }
        self.assertEqual(
            resolve_region("未知", "未知", "另一账号", overrides, author_id="123456"),
            {"city": "深圳", "province": "广东省", "region": "华南", "confidence": "confirmed"},
        )

    def test_dealer_override_beats_city_token_inference(self):
        overrides = {
            "北京经销商": {"city": "长春", "province": "吉林省", "region": "东北"}
        }
        self.assertEqual(
            resolve_region("北京经销商", "上海门店", "未配置账号", overrides),
            {"city": "长春", "province": "吉林省", "region": "东北", "confidence": "confirmed"},
        )

    def test_store_override_beats_city_token_inference(self):
        overrides = {
            "深圳门店": {"city": "成都", "province": "四川省", "region": "西南"}
        }
        self.assertEqual(
            resolve_region("上海经销商", "深圳门店", "未配置账号", overrides),
            {"city": "成都", "province": "四川省", "region": "西南", "confidence": "confirmed"},
        )

    def test_region_resolution_selects_the_longest_city_token(self):
        result = resolve_region("鄂尔多斯上海经销商", "", "", {})
        self.assertEqual(
            result,
            {"city": "鄂尔多斯", "province": "内蒙古自治区", "region": "华北", "confidence": "inferred"},
        )

    def test_local_account_and_store_inference_beats_dealer_legal_name(self):
        result = resolve_region("\u4e0a\u6d77\u7ecf\u9500\u5546\u6709\u9650\u516c\u53f8", "\u82cf\u5dde\u95e8\u5e97", "\u82cf\u5dde\u8d26\u53f7", {})
        self.assertEqual(
            result,
            {"city": "\u82cf\u5dde", "province": "\u6c5f\u82cf\u7701", "region": "\u534e\u4e1c", "confidence": "inferred"},
        )

    def test_unknown_region_is_marked_for_completion(self):
        self.assertEqual(
            resolve_region("无地点", "", "", {}),
            {"city": "待补充区域", "province": "", "region": "", "confidence": "unknown"},
        )

    def test_kpi_author_is_classified_as_core_kpi(self):
        self.assertEqual(classify_account_cohort(" 123456 ", {"123456", "654321"}), "core_kpi")

    def test_non_kpi_author_is_classified_as_expanded_store(self):
        self.assertEqual(classify_account_cohort("999999", {"123456"}), "expanded_store")


class MonthlySourceTests(unittest.TestCase):
    ACCOUNT_SHEET = "\u5c0f\u7ea2\u4e66\u603b\u6570\u636e"
    NOTE_SHEET = "\u5c0f\u7ea2\u4e66\u7b14\u8bb0\u6570\u636e\u5e93"

    def test_a_monthly_source_public_interfaces_are_available(self):
        required = (
            "SourceDiscoveryError",
            "discover_month_sources",
            "read_monthly_accounts",
            "read_notes",
            "read_kpi",
            "profile_quality",
            "validate_month",
            "previous_month",
        )
        self.assertTrue(all(hasattr(build_insight, name) for name in required))

    def make_monthly_workbook(
        self,
        path,
        account_headers=None,
        note_headers=None,
        monthly_metrics=None,
        cumulative_metrics=None,
        source_month="2026-07",
        exported_at="2026-08-01",
        include_note=True,
    ):
        monthly_metrics = monthly_metrics or {
            "reads": 12, "new_fans": 3, "likes": 4, "collects": 5, "comments": 6, "visitors": 7,
        }
        cumulative_metrics = cumulative_metrics or {
            "reads": 120, "fans": 30, "likes": 40, "collects": 50, "comments": 60,
        }
        workbook = Workbook()
        accounts = workbook.active
        accounts.title = self.ACCOUNT_SHEET
        accounts.append(["Monthly source title"])
        accounts.append(account_headers or [
            "\u7ecf\u9500\u5546\u540d\u79f0", "\u95e8\u5e97\u540d\u79f0", "\u5c0f\u7ea2\u4e66\u8d26\u53f7\u540d\u79f0",
            "\u5c0f\u7ea2\u4e66\u4f5c\u8005ID", "\u5c0f\u7ea2\u4e66\u53f7", "\u603b\u6d4f\u89c8", "\u65b0\u589e\u7c89\u4e1d",
            "\u70b9\u8d5e", "\u6536\u85cf", "\u8bc4\u8bba", "\u4e3b\u9875\u8bbf\u5ba2\u6570",
            "\u7b14\u8bb0\u6761\u6570", "\u603b\u6d4f\u89c8", "\u7c89\u4e1d\u6570", "\u70b9\u8d5e", "\u6536\u85cf", "\u8bc4\u8bba",
        ])
        accounts.append([
            "\u5317\u4eac\u7ecf\u9500\u5546", "\u5317\u4eac\u95e8\u5e97", "\u8d26\u53f7 A", "00123", "9988",
            monthly_metrics["reads"], monthly_metrics["new_fans"], monthly_metrics["likes"],
            monthly_metrics["collects"], monthly_metrics["comments"], monthly_metrics["visitors"],
            99, cumulative_metrics["reads"], cumulative_metrics["fans"], cumulative_metrics["likes"],
            cumulative_metrics["collects"], cumulative_metrics["comments"],
        ])

        notes = workbook.create_sheet(self.NOTE_SHEET)
        notes.append(["Notes source title"])
        notes.append(note_headers or [
            "\u7b14\u8bb0\u5f62\u5f0f", "\u7b14\u8bb0\u7c7b\u578b", "\u5bfc\u51fa\u6570\u636e\u65f6\u95f4", "\u7b14\u8bb0ID", "\u4f5c\u8005ID",
            "\u7b14\u8bb0\u53d1\u5e03\u65f6\u95f4", "\u9605\u8bfb\u6b21\u6570", "\u70b9\u8d5e\u6b21\u6570", "\u6536\u85cf\u6b21\u6570", "\u8bc4\u8bba\u6b21\u6570", "\u5206\u4eab\u6b21\u6570", "\u5355\u6761\u7b14\u8bb0\u6da8\u7c89\u6570",
        ])
        if include_note:
            notes.append(["\u56fe\u7247", "\u4ea7\u54c1\u79cd\u8349", exported_at, f"{source_month}-note", "00123", f"{source_month}-15", 12, 3, 4, 5, 6, 7])
        workbook.save(path)

    def make_kpi_workbook(self, path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Q4"
        sheet.append(["\u7ecf\u9500\u5546\u5206\u7ec4", "\u7ecf\u9500\u5546\u5c0f\u7ea2\u4e66\u8d26\u53f7", "\u7cfb\u7edf\u8d26\u53f7\u540d\u79f0", "\u8d26\u53f7ID", "Q4 \u603b\u9605\u8bfb\u91cf\u76ee\u6807", "Q4 \u603b\u4e92\u52a8\u91cf\u76ee\u6807", "Q4 \u603b\u65b0\u589e\u7c89\u4e1d\u76ee\u6807"])
        sheet.append(["\u4e00\u7ec4", "\u8d26\u53f7 A", "\u8d26\u53f7 A", "00123", 300, 90, 12])
        workbook.save(path)

    def test_monthly_reader_uses_labels_and_first_duplicate_measure(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "monthly.xlsx"
            self.make_monthly_workbook(path)

            records = build_insight.read_monthly_accounts(path, self.ACCOUNT_SHEET, "2026-07")

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["author_id"], "00123")
        self.assertEqual(records[0]["reads"], 12.0)
        self.assertEqual(records[0]["likes"], 4.0)

    def test_notes_reader_supports_aliases_and_preserves_identifiers_as_text(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "notes.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = self.NOTE_SHEET
            sheet.append(["Notes title"])
            sheet.append([
                "\u7b14\u8bb0\u7c7b\u578b", "\u7b14\u8bb0\u5f62\u5f0f", "\u65f6\u95f4", "\u7b14\u8bb0ID", "\u4f5c\u8005ID",
                "\u7b14\u8bb0\u53d1\u5e03\u65f6\u95f4", "\u7d2f\u8ba1\u9605\u8bfb\u6570", "\u7d2f\u8ba1\u70b9\u8d5e\u6570", "\u7d2f\u8ba1\u6536\u85cf\u6570", "\u7d2f\u8ba1\u8bc4\u8bba\u6570", "\u7d2f\u8ba1\u5206\u4eab\u6570", "\u7b14\u8bb0\u65b0\u589e\u7c89\u4e1d\u6570",
            ])
            sheet.append(["\u8425\u9500\u6d3b\u52a8", "\u89c6\u9891", "2026-08-01", "000-note", "000-author", "2026-07-31 10:00:00", 12, 3, 4, 5, 6, 7])
            workbook.save(path)

            records = build_insight.read_notes(path, self.NOTE_SHEET, "2026-08-01")

        self.assertEqual(records[0]["note_id"], "000-note")
        self.assertEqual(records[0]["author_id"], "000-author")
        self.assertEqual(records[0]["note_format"], "\u89c6\u9891")
        self.assertEqual(records[0]["category"], "\u8425\u9500\u6d3b\u52a8")
        self.assertEqual(records[0]["publish_date"], "2026-07-31 10:00:00")
        self.assertEqual(records[0]["reads"], 12.0)

    def test_kpi_reader_reads_q4_targets_and_fills_group(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "FY26.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Q4"
            sheet.append(["\u7ecf\u9500\u5546\u5206\u7ec4", "\u7ecf\u9500\u5546\u5c0f\u7ea2\u4e66\u8d26\u53f7", "\u7cfb\u7edf\u8d26\u53f7\u540d\u79f0", "\u8d26\u53f7ID", "Q4 \u603b\u9605\u8bfb\u91cf\u76ee\u6807", "Q4 \u603b\u4e92\u52a8\u91cf\u76ee\u6807", "Q4 \u603b\u65b0\u589e\u7c89\u4e1d\u76ee\u6807"])
            sheet.append(["\u4e00\u7ec4", "Store A", "System A", "000-kpi", 100, 20, 5])
            sheet.append([None, "Store B", "System B", "001-kpi", 200, 40, 10])
            workbook.save(path)

            records = build_insight.read_kpi(path, "Q4")

        self.assertEqual(records[1]["group"], "\u4e00\u7ec4")
        self.assertEqual(records[0]["author_id"], "000-kpi")
        self.assertEqual(records[0]["read_target"], 100.0)
        self.assertEqual(records[0]["interaction_target"], 20.0)
        self.assertEqual(records[0]["fan_target"], 5.0)

    def test_discovery_accepts_one_curated_workbook_and_rejects_zero_or_ambiguous(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            monthly_dir = data_root / "2026-07\u722c\u866b\u6570\u636e"
            monthly_dir.mkdir()
            (data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI").mkdir(parents=True)
            kpi_path = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI" / "FY26.xlsx"
            Workbook().save(kpi_path)
            with self.assertRaises(build_insight.SourceDiscoveryError):
                build_insight.discover_month_sources(data_root, "2026-07")

            self.make_monthly_workbook(monthly_dir / "curated.xlsx")
            sources = build_insight.discover_month_sources(data_root, "2026-07")
            self.assertEqual(sources["monthly_workbook"].name, "curated.xlsx")
            self.assertEqual(sources["kpi_workbook"], kpi_path.resolve())

            self.make_monthly_workbook(monthly_dir / "another.xlsx")
            with self.assertRaisesRegex(build_insight.SourceDiscoveryError, "Ambiguous curated monthly workbooks"):
                build_insight.discover_month_sources(data_root, "2026-07")

    def test_august_build_uses_july_plus_august_monthly_actuals(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            monthly = {
                "2026-07": {"reads": 10, "new_fans": 1, "likes": 1, "collects": 2, "comments": 3, "visitors": 4},
                "2026-08": {"reads": 20, "new_fans": 2, "likes": 2, "collects": 3, "comments": 4, "visitors": 5},
            }
            for month, metrics in monthly.items():
                month_dir = data_root / f"{month}\u722c\u866b\u6570\u636e"
                month_dir.mkdir()
                self.make_monthly_workbook(
                    month_dir / "curated.xlsx",
                    monthly_metrics=metrics,
                    cumulative_metrics={"reads": 9000, "fans": 900, "likes": 900, "collects": 900, "comments": 900},
                    source_month=month,
                )
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            self.make_kpi_workbook(kpi_dir / "FY26.xlsx")
            output_dir = data_root / "output"

            result = subprocess.run(
                [sys.executable, str(SOURCE_ROOT / "scripts" / "build_insight.py"), "--month", "2026-08", "--data-root", str(data_root), "--output-dir", str(output_dir)],
                cwd=SOURCE_ROOT, capture_output=True, text=True, check=False,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            payload = json.loads((output_dir / "insight_data.json").read_text(encoding="utf-8"))
            self.assertEqual(payload["apple"]["network_kpis"]["reads"]["actual"], 30)
            self.assertEqual(payload["apple"]["network_kpis"]["interactions"]["actual"], 15)
            self.assertEqual(payload["apple"]["network_kpis"]["fans"]["actual"], 3)
            self.assertEqual(payload["analysis_rules"]["q4_elapsed_ratio"], 2 / 3)
            self.assertEqual(
                [item["month"] for item in payload["source_files"]["q4_actual_workbooks"]],
                ["2026-07", "2026-08"],
            )

    def test_september_build_uses_all_three_q4_monthly_actuals(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            for index, month in enumerate(("2026-07", "2026-08", "2026-09"), start=1):
                month_dir = data_root / f"{month}\u722c\u866b\u6570\u636e"
                month_dir.mkdir()
                self.make_monthly_workbook(
                    month_dir / "curated.xlsx",
                    monthly_metrics={"reads": index * 10, "new_fans": index, "likes": index, "collects": index + 1, "comments": index + 2, "visitors": index},
                    cumulative_metrics={"reads": 8000, "fans": 800, "likes": 800, "collects": 800, "comments": 800},
                    source_month=month,
                )
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            self.make_kpi_workbook(kpi_dir / "FY26.xlsx")
            output_dir = data_root / "output"

            result = subprocess.run(
                [sys.executable, str(SOURCE_ROOT / "scripts" / "build_insight.py"), "--month", "2026-09", "--data-root", str(data_root), "--output-dir", str(output_dir)],
                cwd=SOURCE_ROOT, capture_output=True, text=True, check=False,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            payload = json.loads((output_dir / "insight_data.json").read_text(encoding="utf-8"))
            self.assertEqual(payload["apple"]["network_kpis"]["reads"]["actual"], 60)
            self.assertEqual(payload["apple"]["network_kpis"]["interactions"]["actual"], 27)
            self.assertEqual(payload["apple"]["network_kpis"]["fans"]["actual"], 6)
            self.assertEqual(payload["analysis_rules"]["q4_elapsed_ratio"], 1)

    def test_missing_prior_q4_month_blocks_publication(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            august_dir = data_root / "2026-08\u722c\u866b\u6570\u636e"
            august_dir.mkdir()
            self.make_monthly_workbook(august_dir / "curated.xlsx", source_month="2026-08")
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            self.make_kpi_workbook(kpi_dir / "FY26.xlsx")
            output_dir = data_root / "output"

            result = subprocess.run(
                [sys.executable, str(SOURCE_ROOT / "scripts" / "build_insight.py"), "--month", "2026-08", "--data-root", str(data_root), "--output-dir", str(output_dir)],
                cwd=SOURCE_ROOT, capture_output=True, text=True, check=False,
            )

            self.assertNotEqual(result.returncode, 0)
            report = json.loads((output_dir / "quality_report.json").read_text(encoding="utf-8"))
            self.assertEqual(report["quality_status"], "failed")
            self.assertTrue(any("2026-07" in error for error in report["errors"]))
            self.assertFalse((output_dir / "insight_data.json").exists())

    def test_month_validation_and_previous_month_cross_january(self):
        self.assertEqual(build_insight.validate_month("2026-07"), "2026-07")
        self.assertEqual(build_insight.previous_month("2026-01"), "2025-12")
        with self.assertRaises(ValueError):
            build_insight.validate_month("2026-7")

    def test_cli_quality_only_build_writes_quality_report(self):
        production_report_path = SOURCE_ROOT / "generated" / "quality_report.json"
        production_report_bytes = production_report_path.read_bytes()
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            output_dir = data_root / "test-output"
            monthly_dir = data_root / "2026-07\u722c\u866b\u6570\u636e"
            monthly_dir.mkdir()
            self.make_monthly_workbook(monthly_dir / "curated.xlsx")
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Q4"
            sheet.append(["\u7ecf\u9500\u5546\u5206\u7ec4", "\u7ecf\u9500\u5546\u5c0f\u7ea2\u4e66\u8d26\u53f7", "\u7cfb\u7edf\u8d26\u53f7\u540d\u79f0", "\u8d26\u53f7ID", "Q4 \u603b\u9605\u8bfb\u91cf\u76ee\u6807", "Q4 \u603b\u4e92\u52a8\u91cf\u76ee\u6807", "Q4 \u603b\u65b0\u589e\u7c89\u4e1d\u76ee\u6807"])
            sheet.append(["\u4e00\u7ec4", "\u8d26\u53f7 A", "\u8d26\u53f7 A", "00123", 100, 20, 5])
            workbook.save(kpi_dir / "FY26.xlsx")

            result = subprocess.run(
                [sys.executable, str(SOURCE_ROOT / "scripts" / "build_insight.py"), "--month", "2026-07", "--data-root", str(data_root), "--output-dir", str(output_dir), "--quality-only"],
                cwd=SOURCE_ROOT,
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr)

            output_path = output_dir / "quality_report.json"
            with output_path.open(encoding="utf-8") as handle:
                report = json.load(handle)
            self.assertEqual(report["source_month"], "2026-07")

        self.assertEqual(production_report_path.read_bytes(), production_report_bytes)

    def test_full_cli_build_writes_both_json_artifacts_without_mutating_production(self):
        production_paths = [SOURCE_ROOT / "generated" / "quality_report.json", SOURCE_ROOT / "generated" / "insight_data.json"]
        production_bytes = {path: path.read_bytes() if path.exists() else None for path in production_paths}
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            output_dir = data_root / "test-output"
            monthly_dir = data_root / "2026-07\u722c\u866b\u6570\u636e"
            monthly_dir.mkdir()
            self.make_monthly_workbook(monthly_dir / "curated.xlsx")
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Q4"
            sheet.append(["\u7ecf\u9500\u5546\u5206\u7ec4", "\u7ecf\u9500\u5546\u5c0f\u7ea2\u4e66\u8d26\u53f7", "\u7cfb\u7edf\u8d26\u53f7\u540d\u79f0", "\u8d26\u53f7ID", "Q4 \u603b\u9605\u8bfb\u91cf\u76ee\u6807", "Q4 \u603b\u4e92\u52a8\u91cf\u76ee\u6807", "Q4 \u603b\u65b0\u589e\u7c89\u4e1d\u76ee\u6807"])
            sheet.append(["\u4e00\u7ec4", "\u8d26\u53f7 A", "\u8d26\u53f7 A", "00123", 100, 20, 5])
            workbook.save(kpi_dir / "FY26.xlsx")

            result = subprocess.run(
                [sys.executable, str(SOURCE_ROOT / "scripts" / "build_insight.py"), "--month", "2026-07", "--data-root", str(data_root), "--output-dir", str(output_dir)],
                cwd=SOURCE_ROOT,
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            with (output_dir / "quality_report.json").open(encoding="utf-8") as handle:
                self.assertEqual(json.load(handle)["source_month"], "2026-07")
            with (output_dir / "insight_data.json").open(encoding="utf-8") as handle:
                payload = json.load(handle)
                self.assertEqual(payload["schema_version"], "2.0")
                self.assertIn("period", payload)
                self.assertIn("ai_insights", payload["apple"])
            self.assertTrue((output_dir / "month_index.json").exists())
            self.assertTrue((output_dir / "months" / "2026-07" / "apple.json").exists())
            self.assertTrue((output_dir / "dealer_index.json").exists())

        self.assertEqual({path: path.read_bytes() if path.exists() else None for path in production_paths}, production_bytes)

    def test_cli_accepts_ai_and_versioned_month_output(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            output_dir = data_root / "test-output"
            monthly_dir = data_root / "2026-07\u722c\u866b\u6570\u636e"
            monthly_dir.mkdir()
            self.make_monthly_workbook(monthly_dir / "curated.xlsx")
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            self.make_kpi_workbook(kpi_dir / "FY26.xlsx")

            result = build_insight.main([
                "--month", "2026-07",
                "--data-root", str(data_root),
                "--output-dir", str(output_dir),
                "--ai",
            ])

            self.assertEqual(result, 0)
            self.assertTrue((output_dir / "month_index.json").exists())
            self.assertTrue((output_dir / "months" / "2026-07" / "apple.json").exists())
            self.assertTrue((output_dir / "insight_data.json").exists())
            self.assertTrue((output_dir / "dealer_index.json").exists())

    def test_cli_no_compat_skips_latest_paths(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            output_dir = data_root / "test-output"
            monthly_dir = data_root / "2026-07\u722c\u866b\u6570\u636e"
            monthly_dir.mkdir()
            self.make_monthly_workbook(monthly_dir / "curated.xlsx")
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            self.make_kpi_workbook(kpi_dir / "FY26.xlsx")

            result = build_insight.main([
                "--month", "2026-07",
                "--data-root", str(data_root),
                "--output-dir", str(output_dir),
                "--no-compat",
            ])

            self.assertEqual(result, 0)
            self.assertTrue((output_dir / "months" / "2026-07" / "apple.json").exists())
            self.assertFalse((output_dir / "insight_data.json").exists())
            self.assertFalse((output_dir / "dealer_index.json").exists())

    def test_failed_quality_writes_report_returns_nonzero_and_does_not_publish_insight(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            monthly_dir = data_root / "2026-07\u722c\u866b\u6570\u636e"
            monthly_dir.mkdir()
            self.make_monthly_workbook(monthly_dir / "curated.xlsx", include_note=False)
            kpi_dir = data_root / "\u5386\u53f2\u6570\u636e\u5bfc\u5165" / "\u7ecf\u9500\u5546KPI"
            kpi_dir.mkdir(parents=True)
            self.make_kpi_workbook(kpi_dir / "FY26.xlsx")
            output_dir = data_root / "isolated-output"

            result = subprocess.run(
                [sys.executable, str(SOURCE_ROOT / "scripts" / "build_insight.py"), "--month", "2026-07", "--data-root", str(data_root), "--output-dir", str(output_dir)],
                cwd=SOURCE_ROOT, capture_output=True, text=True, check=False,
            )

            self.assertNotEqual(result.returncode, 0)
            report = json.loads((output_dir / "quality_report.json").read_text(encoding="utf-8"))
            self.assertEqual(report["quality_status"], "failed")
            self.assertTrue(report["errors"])
            self.assertFalse((output_dir / "insight_data.json").exists())

    def test_payload_builder_rejects_failed_quality_contract(self):
        quality = {"quality_status": "failed", "errors": ["No note rows"], "source_files": {}}
        with self.assertRaisesRegex(ValueError, "not publishable"):
            build_insight.build_insight_payload([], [], [], {}, {}, quality, "2026-07")

    def test_quality_profile_reports_grain_completeness_matching_and_warnings(self):
        accounts = [
            {"author_id": "a1", "account_name": "\u5317\u4eac\u8d26\u53f7", "dealer": "", "store": ""},
            {"author_id": "a2", "account_name": "\u672a\u77e5\u8d26\u53f7", "dealer": "", "store": ""},
        ]
        notes = [
            {"note_id": "n1", "author_id": "a1", "category": "\u4ea7\u54c1\u79cd\u8349"},
            {"note_id": "n1", "author_id": "a1", "category": ""},
            {"note_id": "n2", "author_id": "a2", "category": ""},
        ]
        kpis = [
            {"author_id": "a1", "account_name": "Known KPI"},
            {"author_id": "missing", "account_name": "Missing KPI"},
        ]

        report = build_insight.profile_quality(accounts, notes, kpis, "2026-07", {"monthly_workbook": Path("monthly.xlsx")})

        self.assertEqual(report["previous_month"], "2026-06")
        self.assertEqual(report["account_rows"], 2)
        self.assertEqual(report["unique_author_ids"], 2)
        self.assertEqual(report["note_rows"], 3)
        self.assertEqual(report["duplicate_note_ids"], ["n1"])
        self.assertEqual(report["kpi_accounts"], 2)
        self.assertEqual(report["matched_kpi_accounts"], 1)
        self.assertEqual(report["kpi_match_rate"], 0.5)
        self.assertEqual(report["category_completeness"]["missing"], 2)
        self.assertEqual(report["city_identification"]["rate"], 0.5)
        self.assertEqual(report["account_cohorts"], {"core_kpi": 1, "expanded_store": 1})
        self.assertEqual(report["quality_status"], "ready_with_warnings")
        self.assertEqual(
            {warning["code"] for warning in report["warnings"]},
            {"scope_change", "duplicate_note_ids", "missing_categories", "unknown_cities", "unmatched_kpi_ids"},
        )

    def test_data_freshness_uses_note_export_time_and_is_distinct_from_build_time(self):
        accounts = [{
            "author_id": "00123", "account_name": "Old Account", "dealer": "Old Dealer", "store": "",
            "reads": 10, "new_fans": 1, "likes": 1, "collects": 1, "comments": 1,
        }]
        notes = [{
            "note_id": "old-note", "author_id": "00123", "category": "A", "note_format": "image",
            "publish_date": "2024-01-15", "exported_at": "2024-02-01", "reads": 10,
            "likes": 1, "collects": 1, "comments": 1, "new_fans": 1,
        }]
        kpis = [{"author_id": "00123", "read_target": 100, "interaction_target": 10, "fan_target": 2}]
        report = build_insight.profile_quality(
            accounts, notes, kpis, "2024-01", {"monthly_workbook": Path("old-month.xlsx")},
        )

        self.assertEqual(report.get("data_freshness"), {
            "source_snapshot_at": "2024-02-01",
            "basis": "note_export_timestamp",
            "is_fallback": False,
        })
        self.assertNotEqual(report["data_freshness"]["source_snapshot_at"], report["generated_at"][:10])

        payload = build_insight.build_insight_payload(
            accounts, notes, kpis, {"A": {"unified": "A", "confirmed": True}}, {}, report, "2024-01",
        )
        self.assertEqual(payload["metadata"]["data_freshness"], report["data_freshness"])
        self.assertEqual(payload["metadata"]["generated_at"], payload["generated_at"])
        self.assertEqual(payload["apple"]["quality_metadata"]["data_freshness"], report["data_freshness"])

    def test_data_freshness_falls_back_to_source_month_end_when_export_time_missing(self):
        report = build_insight.profile_quality(
            [{"author_id": "a", "account_name": "A", "dealer": "A", "store": ""}],
            [{"note_id": "n", "author_id": "a", "category": "A", "exported_at": ""}],
            [{"author_id": "a", "account_name": "A"}],
            "2026-07",
            {"monthly_workbook": Path("monthly.xlsx")},
        )

        self.assertEqual(report.get("data_freshness"), {
            "source_snapshot_at": "2026-07-31",
            "basis": "source_month_end_fallback",
            "is_fallback": True,
        })
