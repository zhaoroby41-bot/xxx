import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


SOURCE_ROOT = Path(__file__).resolve().parents[1]
if str(SOURCE_ROOT) not in sys.path:
    sys.path.insert(0, str(SOURCE_ROOT))

from scripts.fiscal_history import (
    apple_period,
    build_history_context,
    discover_history_workbooks,
)


class FiscalHistoryTests(unittest.TestCase):
    def test_october_starts_next_apple_fiscal_year(self):
        self.assertEqual(
            apple_period("2025-10"),
            {"calendar_month": "2025-10", "fiscal_year": "FY26", "fiscal_quarter": "Q1", "fiscal_month": 1},
        )

    def test_july_is_q4_month_one(self):
        period = apple_period("2026-07")
        self.assertEqual((period["fiscal_year"], period["fiscal_quarter"], period["fiscal_month"]), ("FY26", "Q4", 1))

    def test_history_discovery_is_month_sorted_and_stops_before_current(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            for name in ("小红书总数据2026-06.xlsx", "小红书总数据2025 - 7.xlsx", "小红书总数据2026年7月.xlsx"):
                Workbook().save(root / name)
            sources = discover_history_workbooks(root, "2026-07")
            self.assertEqual([item.month for item in sources], ["2025-07", "2026-06"])

    def test_history_discovery_rejects_duplicate_months(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            Workbook().save(root / "总数据2026 - 6.xlsx")
            Workbook().save(root / "总数据2026-06.xlsx")
            with self.assertRaises(ValueError):
                discover_history_workbooks(root, "2026-07")

    def test_history_context_returns_previous_yoy_and_rolling_median(self):
        result = build_history_context(
            {"month": "2026-07", "reads": 130, "notes": 10},
            [
                {"month": "2025-07", "reads": 100, "notes": 8},
                {"month": "2026-05", "reads": 90, "notes": 9},
                {"month": "2026-06", "reads": 120, "notes": 10},
            ],
        )
        self.assertEqual(result["previous_month"]["month"], "2026-06")
        self.assertEqual(result["year_ago"]["month"], "2025-07")
        self.assertEqual(result["rolling_baseline"]["reads"], 100)
        self.assertEqual(result["comparisons"]["previous_month"]["reads"]["absolute_change"], 10)
        self.assertAlmostEqual(result["comparisons"]["year_ago"]["reads"]["ratio_change"], 0.3)

    def test_history_context_does_not_skip_a_missing_previous_month(self):
        result = build_history_context(
            {"month": "2026-08", "reads": 130},
            [{"month": "2026-06", "reads": 120}],
        )
        self.assertIsNone(result["previous_month"])
        self.assertIsNone(result["comparisons"]["previous_month"]["reads"]["absolute_change"])

    def test_missing_history_metrics_remain_null(self):
        result = build_history_context({"month": "2026-07", "reads": 20}, [])
        self.assertIsNone(result["previous_month"])
        self.assertIsNone(result["rolling_baseline"]["reads"])
        self.assertIsNone(result["comparisons"]["previous_month"]["reads"]["absolute_change"])
        self.assertIsNone(result["comparisons"]["rolling_baseline"]["reads"]["ratio_change"])


if __name__ == "__main__":
    unittest.main()
